"""Merge per-lesson/assignment remarks + root Extra.xlsx into overview.json."""
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from utils.anonymize import is_llm_category, load_student_category_ids
from utils.folder_utils import (
    find_working_remarks, normalize_sid as _normalize_sid, pick_folder)

try:
    from openpyxl import load_workbook
except ImportError:
    print('error: openpyxl required. pip install openpyxl', file=sys.stderr)
    sys.exit(1)


_LESSON_ORDER_FILE = 'lesson_order.txt'

_BLOCK_COLS = (
    'HTML Follow', 'CSS Follow', 'JS Follow', 'Follow', 'Inc',
    'A', 'Q', 'H', 'C+', 'C-', 'C Diff',
    'LessonObs', 'Grade', 'Status', 'Obs',
)

_BLOCK_FIELD_ROLE = {
    'HTML Follow': 'follow_html', 'CSS Follow': 'follow_css',
    'JS Follow': 'follow_js', 'Follow': 'follow', 'Inc': 'inc',
    'A': 'a', 'Q': 'q', 'H': 'h', 'C+': 'c_plus', 'C-': 'c_minus',
    'C Diff': 'c_diff', 'LessonObs': 'lesson_obs', 'Grade': 'grade',
    'Status': 'status', 'Obs': 'obs',
}

_ROLE_ALIASES = {
    'id':              ('ID',),
    'name':            ('Name',),
    'number':          ('Number',),
    'excluded':        ('Category',),
    'final_grade':     ('Final Grade', 'Grade'),
    'avg_assignments': ('Avg Assignments', 'Avg Grade', 'Average'),
    'participation':   ('Participation',),
    'pre_typing':      ('Pre K/min',),
    'post_typing':     ('Post K/min',),
    'self_eval':       ('Self Eval', 'Self Evaluation', 'Self'),
    'kahoot':          ('Kahoot',),
    'quiz_stii':       ('Final Quiz', 'Quiz Stii', 'Stii', 'Știi'),
    'answers':         ('Total Answers', 'Answers'),
    'questions':       ('Total Questions', 'Questions'),
    'help':            ('Total Help', 'Help'),
}

_EXTRA_IDENTITY = {'id', 'name', 'number'}


def _build_columns_contract(header: Sequence[str],
                            topics: Sequence[str]) -> Dict[str, object]:
    lower_to_idx: Dict[str, int] = {}
    for i, h in enumerate(header):
        key = str(h).strip().lower()
        if key and key not in lower_to_idx:
            lower_to_idx[key] = i
    roles: Dict[str, int] = {}
    for role, aliases in _ROLE_ALIASES.items():
        for a in aliases:
            idx = lower_to_idx.get(a.lower())
            if idx is not None:
                roles[role] = idx
                break
    topic_list = []
    for lk in topics:
        prefix = lk.title()
        fields: Dict[str, int] = {}
        for bc in _BLOCK_COLS:
            idx = lower_to_idx.get(f'{prefix} {bc}'.lower())
            if idx is not None:
                fields[_BLOCK_FIELD_ROLE[bc]] = idx
        topic_list.append({'name': lk, 'label': prefix, 'fields': fields})
    return {'roles': roles, 'topics': topic_list}


def _read_lesson_order(root: Optional[Path]) -> list:
    if root is None:
        return []
    try:
        lines = (root / _LESSON_ORDER_FILE).read_text(
            encoding='utf-8-sig', errors='replace').splitlines()
    except (OSError, ValueError):
        return []
    order: list = []
    for line in lines:
        name = line.split('#', 1)[0].strip().lower()
        if name and name not in order:
            order.append(name)
    return order


def _ordered_topics(lessons_root: Optional[Path],
                    assignments_root: Optional[Path],
                    order: Sequence[str]) -> list:
    names: set = set()
    for group_root in (lessons_root, assignments_root):
        if group_root is not None and group_root.is_dir():
            for d in group_root.iterdir():
                if d.is_dir():
                    names.add(d.name.lower())
    known = [k for k in order if k in names]
    extras = sorted(k for k in names if k not in order)
    return known + extras


def _find_subdir(parent: Path, name: str) -> Optional[Path]:
    if not parent.is_dir():
        return None
    for entry in parent.iterdir():
        if entry.is_dir() and entry.name.lower() == name.lower():
            return entry
    return None


def _read_grades_rows(xlsx_path: Path,
                      wanted_headers: Tuple[str, ...]
                      ) -> Tuple[Dict[str, int], list]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {}, []
    header = [str(c).strip() if c is not None else '' for c in rows[0]]
    col_idx = {}
    for h in wanted_headers:
        try:
            col_idx[h] = header.index(h)
        except ValueError:
            col_idx[h] = None
    out_rows = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        d = {}
        for h, idx in col_idx.items():
            d[h] = r[idx] if idx is not None and idx < len(r) else None
        out_rows.append(d)
    return col_idx, out_rows


def _to_str(v) -> str:
    if v is None:
        return ''
    return str(v).strip()


def _count_interactions(raw) -> Tuple[int, int, int]:
    s = _to_str(raw)
    if not s:
        return (0, 0, 0)
    parts = [p.strip() for p in s.split(',') if p.strip()]
    a = sum(1 for p in parts if p == 'A')
    q = sum(1 for p in parts if p == 'Q')
    h = sum(1 for p in parts if p == 'H')
    return (a, q, h)


def _count_extra_missing(desc) -> Tuple[int, int]:
    s = _to_str(desc)
    if not s:
        return (0, 0)
    plus, minus = 0, 0
    for item in s.split(','):
        item = item.strip()
        if not item or item[0] not in '+-':
            continue
        n = 1
        if item.endswith(')') and ' (x' in item:
            try:
                n = int(item[item.rindex('(x') + 2 : -1])
            except (ValueError, IndexError):
                n = 1
        if item[0] == '+':
            plus += n
        else:
            minus += n
    return (plus, minus)


def _read_lesson_stats_csv(csv_path: Path) -> Optional[Dict[str, object]]:
    if not csv_path.is_file():
        return None
    try:
        text = csv_path.read_text(encoding='utf-8')
    except OSError:
        return None
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    if len(lines) < 2:
        return None
    header = [c.strip() for c in lines[0].split(',')]
    values = [c.strip() for c in lines[1].split(',')]
    out: Dict[str, object] = {}
    for k, v in zip(header, values):
        if not k:
            continue
        if v == '':
            out[k] = ''
            continue
        try:
            if '.' in v:
                out[k] = float(v)
            else:
                out[k] = int(v)
        except ValueError:
            out[k] = v
    return out


_LESSON_STATS_SOURCES = ('lesson_stats_py.csv', 'lesson_stats.csv')


def _collect_lesson_stats_rows(lessons_root: Optional[Path],
                               topics: Sequence[str]
                               ) -> Optional[Dict[str, list]]:
    if lessons_root is None or not lessons_root.is_dir():
        return None
    rows: list = []
    column_order: list = []
    seen = set()
    dirs_by_lower = {
        d.name.lower(): d
        for d in lessons_root.iterdir()
        if d.is_dir()
    }
    ordered_keys = [lk for lk in topics if lk in dirs_by_lower]
    ordered_keys += sorted(lk for lk in dirs_by_lower if lk not in topics)
    for lk in ordered_keys:
        lesson_dir = dirs_by_lower[lk]
        data = None
        for fname in _LESSON_STATS_SOURCES:
            data = _read_lesson_stats_csv(lesson_dir / fname)
            if data is not None:
                break
        if data is None:
            continue
        rows.append((lk.title(), data))
        for k in data.keys():
            if k not in seen:
                seen.add(k)
                column_order.append(k)
    if not rows:
        return None
    header = ['Lesson'] + column_order
    out_rows = [[name] + [data.get(k, '') for k in column_order]
                for name, data in rows]
    return {'header': header, 'rows': out_rows}


def _update_identity(student_data: Dict[str, dict], sid: str, r: dict) -> None:
    name = _to_str(r.get('Student'))
    number = _to_str(r.get('Number'))
    cat = _to_str(r.get('Category'))
    if name and 'name' not in student_data[sid]:
        student_data[sid]['name'] = name
    if number and 'number' not in student_data[sid]:
        student_data[sid]['number'] = number
    if cat and 'category' not in student_data[sid]:
        student_data[sid]['category'] = cat


def _collect_student_data(lessons_root: Optional[Path],
                          assignments_root: Optional[Path],
                          topics: set,
                          excluded_ids) -> Tuple[Dict[str, dict], Dict[str, dict]]:
    student_data: Dict[str, dict] = defaultdict(dict)
    lesson_meta: Dict[str, dict] = {}

    if lessons_root is not None:
        for lesson_dir in sorted(d for d in lessons_root.iterdir() if d.is_dir()):
            lesson_key = lesson_dir.name.lower()
            if lesson_key not in topics:
                continue
            xlsx = find_working_remarks(lesson_dir)
            if xlsx is None:
                print(f'  [lesson/{lesson_dir.name}] skipped: no remarks.xlsx or remarks_<ts>.xlsx')
                continue
            wanted = ('ID', 'Student', 'Number', 'Category', 'Inc',
                      'Interactions',
                      'Follow (C) Desc', 'Follow (E)',
                      'HTML (E)', 'CSS (E)', 'JS (E)', 'Obs')
            col_idx, rows = _read_grades_rows(xlsx, wanted)
            missing = [h for h in wanted if col_idx.get(h) is None]
            for h in ('Inc', 'Interactions', 'Follow (E)', 'Obs'):
                if h in missing:
                    print(f'  [lesson/{lesson_dir.name}] WARNING: column '
                          f'{h!r} not found in {xlsx.name}')
            lesson_meta[lesson_key] = {
                'has_interactions':  col_idx.get('Interactions') is not None,
                'has_follow_c_desc': col_idx.get('Follow (C) Desc') is not None,
            }
            n = 0
            for r in rows:
                sid = _normalize_sid(r.get('ID'))
                if not sid:
                    continue
                _update_identity(student_data, sid, r)
                if sid in excluded_ids:
                    continue
                student_data[sid][f'lesson_{lesson_key}'] = {
                    'follow':         r.get('Follow (E)'),
                    'follow_html':    r.get('HTML (E)'),
                    'follow_css':     r.get('CSS (E)'),
                    'follow_js':      r.get('JS (E)'),
                    'inc':            r.get('Inc'),
                    'interactions':   _to_str(r.get('Interactions')),
                    'follow_c_desc':  _to_str(r.get('Follow (C) Desc')),
                    'obs':            _to_str(r.get('Obs')),
                }
                n += 1
            print(f'  [lesson/{lesson_dir.name}] {n} student row(s) from {xlsx.name}')

    if assignments_root is not None:
        for assign_dir in sorted(d for d in assignments_root.iterdir() if d.is_dir()):
            assign_key = assign_dir.name.lower()
            if assign_key not in topics:
                continue
            xlsx = find_working_remarks(assign_dir)
            if xlsx is None:
                print(f'  [assign/{assign_dir.name}] skipped: no remarks.xlsx or remarks_<ts>.xlsx')
                continue
            wanted = ('ID', 'Student', 'Number', 'Category', 'Grade', 'Status', 'Obs')
            col_idx, rows = _read_grades_rows(xlsx, wanted)
            missing = [h for h in wanted if col_idx.get(h) is None]
            for h in ('Grade', 'Status', 'Obs'):
                if h in missing:
                    print(f'  [assign/{assign_dir.name}] WARNING: column '
                          f'{h!r} not found in {xlsx.name}')
            n = 0
            for r in rows:
                sid = _normalize_sid(r.get('ID'))
                if not sid:
                    continue
                _update_identity(student_data, sid, r)
                if sid in excluded_ids:
                    continue
                student_data[sid][f'assign_{assign_key}'] = {
                    'grade':  r.get('Grade'),
                    'status': _to_str(r.get('Status')),
                    'obs':    _to_str(r.get('Obs')),
                }
                n += 1
            print(f'  [assign/{assign_dir.name}] {n} student row(s) from {xlsx.name}')

    return student_data, lesson_meta


def _build_row_values(sid: str,
                      info: dict,
                      lesson_meta: Dict[str, dict],
                      topics: Sequence[str],
                      excluded: bool,
                      is_llm: bool = False) -> Dict[str, Any]:
    category = 'EXCLUDED' if excluded else ('LLM' if is_llm else None)
    out: Dict[str, Any] = {
        'ID': sid,
        'Name': info.get('name', ''),
        'Number': info.get('number', ''),
        'Category': category,
    }
    if excluded:
        return out
    for lk in topics:
        prefix = lk.title()
        lesson_info = info.get(f'lesson_{lk}')
        meta = lesson_meta.get(lk, {})
        if lesson_info:
            out[f'{prefix} HTML Follow'] = lesson_info.get('follow_html')
            out[f'{prefix} CSS Follow'] = lesson_info.get('follow_css')
            out[f'{prefix} JS Follow'] = lesson_info.get('follow_js')
            out[f'{prefix} Follow'] = lesson_info.get('follow')
            out[f'{prefix} Inc'] = lesson_info.get('inc')
            attended = lesson_info.get('follow') not in (None, '')
            if attended and meta.get('has_interactions'):
                a, q, h = _count_interactions(lesson_info.get('interactions'))
                out[f'{prefix} A'] = a
                out[f'{prefix} Q'] = q
                out[f'{prefix} H'] = h
            if attended and meta.get('has_follow_c_desc'):
                cplus, cminus = _count_extra_missing(lesson_info.get('follow_c_desc'))
                out[f'{prefix} C+'] = cplus
                out[f'{prefix} C-'] = cminus
                out[f'{prefix} C Diff'] = cplus - cminus
            out[f'{prefix} LessonObs'] = lesson_info.get('obs', '')
        assign_info = info.get(f'assign_{lk}')
        if assign_info:
            out[f'{prefix} Grade'] = assign_info.get('grade')
            out[f'{prefix} Status'] = assign_info.get('status', '')
            out[f'{prefix} Obs'] = assign_info.get('obs', '')
    return out


def _read_extra_xlsx(root: Path) -> Tuple[
        Dict[str, dict], List[str], List[str], List[Tuple[str, str]]]:
    empty = ({}, [], [], [])
    path = root / 'Extra.xlsx'
    if not path.is_file():
        return empty
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb['Grades'] if 'Grades' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return empty

    header = [_to_str(c) for c in rows[0]]
    id_idx = next((i for i, h in enumerate(header) if h.lower() == 'id'), None)
    if id_idx is None:
        print('  [Extra.xlsx] WARNING: no ID column; skipped')
        return empty

    sep_idx = next((i for i, h in enumerate(header) if h == ''), len(header))

    def _is_data_col(i: int) -> bool:
        return bool(header[i]) and header[i].lower() not in _EXTRA_IDENTITY

    before = [i for i in range(len(header)) if i < sep_idx and _is_data_col(i)]
    after = [i for i in range(len(header)) if i > sep_idx and _is_data_col(i)]
    dup = ({header[i].lower() for i in before}
           & {header[i].lower() for i in after})

    def _canonical(i: int, is_after: bool) -> str:
        name = header[i]
        if name.lower() in dup:
            return ('Post ' if is_after else 'Pre ') + name
        return name

    before_spec = [(i, _canonical(i, False)) for i in before]
    after_spec = [(i, _canonical(i, True)) for i in after]
    before_names = [name for _, name in before_spec]
    after_names = [name for _, name in after_spec]
    pairs = [('Pre ' + header[i], 'Post ' + header[i])
             for i in before if header[i].lower() in dup]

    extra_by_sid: Dict[str, dict] = {}
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        sid = _normalize_sid(r[id_idx]) if id_idx < len(r) else ''
        if not sid:
            continue
        extra_by_sid[sid] = {
            name: (r[i] if i < len(r) else None)
            for i, name in before_spec + after_spec
        }
    print(f'  [Extra.xlsx] {len(extra_by_sid)} row(s); '
          f'before: {", ".join(before_names) or "(none)"}; '
          f'after: {", ".join(after_names) or "(none)"}')
    return extra_by_sid, before_names, after_names, pairs


def _sid_sort_key(sid: str):
    try:
        return (0, int(sid))
    except (ValueError, TypeError):
        return (1, sid)


def _build_table(student_data: Dict[str, dict],
                 lesson_meta: Dict[str, dict],
                 topics: Sequence[str],
                 excluded_ids,
                 llm_ids,
                 extra_by_sid: Dict[str, dict],
                 extra_before: List[str],
                 extra_after: List[str]) -> Tuple[List[str], List[list]]:
    header: List[str] = ['ID', 'Name', 'Number', 'Category']
    header += extra_before
    for lk in topics:
        prefix = lk.title()
        header += [f'{prefix} {bc}' for bc in _BLOCK_COLS]
    header += extra_after

    rows: List[list] = []
    for sid in sorted(student_data.keys(), key=_sid_sort_key):
        info = student_data[sid]
        d = _build_row_values(
            sid, info, lesson_meta, topics,
            sid in excluded_ids, sid in llm_ids,
        )
        ex = extra_by_sid.get(sid, {})
        row = [d[h] if h in d else (ex[h] if h in ex else None)
               for h in header]
        rows.append(row)
    return header, rows


def main(argv) -> int:
    parser = argparse.ArgumentParser(
        description='Merge per-lesson/assignment remarks + Extra.xlsx into '
                    'overview.json.')
    parser.add_argument('root', nargs='?',
                        help='Course root folder (containing lessons/ and assignments/).')
    parser.add_argument('--output', default=None,
                        help='Output filename (relative to root). '
                             'Defaults to overview.json.')
    parser.add_argument('--no-stats', action='store_true',
                        help='Skip the chained analyze_grades step '
                             '(grades_stats.json will not be refreshed).')
    args = parser.parse_args(argv[1:])

    if args.root:
        root = Path(args.root)
    else:
        chosen = pick_folder('Select the course root folder')
        root = Path(chosen) if chosen else None
    if root is None or not root.is_dir():
        print('No root folder selected.', file=sys.stderr)
        return 1
    root = root.resolve()

    lessons_root = _find_subdir(root, 'lessons')
    assignments_root = _find_subdir(root, 'assignments')
    if lessons_root is None and assignments_root is None:
        print(f'error: no lessons/ or assignments/ folder in {root}',
              file=sys.stderr)
        return 1

    topics = _ordered_topics(
        lessons_root, assignments_root, _read_lesson_order(root))
    topics_set = set(topics)

    print(f'Root:        {root}')
    print(f'Lessons:     {lessons_root or "(missing)"}')
    print(f'Assignments: {assignments_root or "(missing)"}')
    print()

    excluded_ids, llm_ids = load_student_category_ids(str(root / 'students.csv'))
    if excluded_ids:
        print(f'Excluded:    {len(excluded_ids)} student(s) per '
              f'students.csv (Category=Excluded)')
    if llm_ids:
        print(f'LLM/AI:      {len(llm_ids)} row(s) per '
              f'students.csv (Category=LLM)')
    if excluded_ids or llm_ids:
        print()

    student_data, lesson_meta = _collect_student_data(
        lessons_root, assignments_root, topics_set, excluded_ids,
    )

    if not student_data:
        print('\nNo student rows collected.', file=sys.stderr)
        return 1

    for sid, info in student_data.items():
        if is_llm_category(info.get('category')):
            llm_ids.add(sid)
    if llm_ids:
        print(f'LLM/AI total: {len(llm_ids)} row(s) '
              f'(students.csv + remarks sheets)')

    extra_by_sid, extra_before, extra_after, extra_pairs = _read_extra_xlsx(root)

    header, rows = _build_table(
        student_data, lesson_meta, topics, excluded_ids, llm_ids,
        extra_by_sid, extra_before, extra_after,
    )
    lesson_stats = _collect_lesson_stats_rows(lessons_root, topics)

    out_path = root / (args.output or 'overview.json')
    payload = {'header': header, 'rows': rows}
    payload['columns'] = _build_columns_contract(header, topics)
    if extra_before or extra_after:
        payload['extra_columns'] = {
            'before': extra_before,
            'after': extra_after,
            'pairs': [list(p) for p in extra_pairs],
        }
    if lesson_stats:
        payload['lesson_stats'] = lesson_stats
    out_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=1), encoding='utf-8')

    print(f'\nWrote {out_path}')
    print(f'  {len(rows)} student row(s)')

    if not args.no_stats:
        _run_analyze_grades(out_path)

    return 0


def _run_analyze_grades(json_path: Path) -> None:
    script = Path(__file__).resolve().parent / 'analyze_grades.py'
    print(f'\nGenerating grades_stats.json…')
    try:
        subprocess.run(
            [sys.executable, str(script), str(json_path), '--no-plot'],
            check=True,
        )
    except (subprocess.CalledProcessError, OSError) as exc:
        print(f'  WARNING: analyze_grades step failed ({exc}); '
              f'grades_stats.json not refreshed', file=sys.stderr)


if __name__ == '__main__':
    sys.exit(main(sys.argv))
