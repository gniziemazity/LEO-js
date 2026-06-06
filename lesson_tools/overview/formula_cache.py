from __future__ import annotations

import re
import shutil
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, Optional, Tuple

from openpyxl import load_workbook

_XL_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'


def _snapshot_cached_values(path: Path) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    if not path or not path.exists():
        return out
    wb = load_workbook(path, data_only=True)
    try:
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            sheet_cache: Dict[str, Any] = {}
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    if cell.data_type == 'n':
                        sheet_cache[cell.coordinate] = cell.value
            if sheet_cache:
                out[ws_name] = sheet_cache
    finally:
        wb.close()
    return out


def _normalize_target_path(target: str) -> str:
    target = target.lstrip('/')
    if target.startswith('xl/'):
        return target
    return 'xl/' + target


_CELL_REF_RE = re.compile(r'\$?[A-Z]+\$?\d+')


def _resolve_ref(ws, ref: str):
    return ws[ref.replace('$', '')].value


def _split_top_level_args(args_str: str):
    args = []
    depth = 0
    in_str = False
    last = 0
    for i, ch in enumerate(args_str):
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == '(':
            depth += 1
        elif ch == ')':
            depth -= 1
        elif ch == ',' and depth == 0:
            args.append(args_str[last:i].strip())
            last = i + 1
    args.append(args_str[last:].strip())
    return args


def _excel_if_to_python(expr: str):
    while True:
        i = expr.find('IF(')
        if i < 0:
            return expr
        # find matching close paren for this IF(
        depth = 0
        end = -1
        in_str = False
        for j in range(i + 3, len(expr)):
            ch = expr[j]
            if ch == '"':
                in_str = not in_str
                continue
            if in_str:
                continue
            if ch == '(':
                depth += 1
            elif ch == ')':
                if depth == 0:
                    end = j
                    break
                depth -= 1
        if end < 0:
            return None
        args = _split_top_level_args(expr[i + 3:end])
        if len(args) != 3:
            return None
        cond, a, b = args
        cond_py = re.sub(r'(?<![<>!=])=(?!=)', '==', cond)
        cond_py = cond_py.replace('<>', '!=')
        replacement = f'(({a}) if ({cond_py}) else ({b}))'
        expr = expr[:i] + replacement + expr[end + 1:]


def _eval_formula(formula_text: str, ws) -> Tuple[Optional[Any], bool]:
    if not isinstance(formula_text, str):
        return None, False
    expr = formula_text.lstrip('=').strip()
    if not expr:
        return None, False

    refs = sorted(set(_CELL_REF_RE.findall(expr)), key=len, reverse=True)
    for ref in refs:
        try:
            val = _resolve_ref(ws, ref)
        except Exception:
            return None, False
        if isinstance(val, str) and val.startswith('='):
            return None, False  # depends on a sibling formula
        if val is None:
            return None, False  # missing input → skip, can't cache
        if isinstance(val, (int, float)):
            replacement = repr(val)
        elif isinstance(val, str):
            replacement = repr(val)
        else:
            return None, False
        expr = re.sub(r'\$?' + re.escape(ref.replace('$', '')) + r'\b',
                      replacement, expr)

    expr = _excel_if_to_python(expr)
    if expr is None:
        return None, False
    expr = expr.replace('<>', '!=')

    _ALLOWED_KEYWORDS = {'if', 'else', 'and', 'or', 'not', 'True', 'False'}
    for m in re.finditer(r'\b([A-Za-z_][A-Za-z_0-9]*)\s*\(', expr):
        if m.group(1) not in _ALLOWED_KEYWORDS:
            return None, False

    try:
        value = eval(expr, {'__builtins__': {}}, {})
    except ZeroDivisionError:
        return None, False
    except Exception:
        return None, False
    if isinstance(value, bool):
        return value, True
    if isinstance(value, (int, float)):
        return value, True
    return None, False


def _evaluate_formulas_into_snapshot(path: Path,
                                     base: Dict[str, Dict[str, Any]]
                                     ) -> Dict[str, Dict[str, Any]]:
    wb = load_workbook(path)
    try:
        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not isinstance(v, str) or not v.startswith('='):
                        continue
                    sheet_cache = base.setdefault(ws_name, {})
                    if cell.coordinate in sheet_cache:
                        continue
                    val, ok = _eval_formula(v, ws)
                    if ok:
                        sheet_cache[cell.coordinate] = val
    finally:
        wb.close()
    return base


def _patch_cached_values(path: Path,
                         cache: Dict[str, Dict[str, Any]]) -> int:
    if not cache:
        return 0

    with zipfile.ZipFile(path, 'r') as z:
        try:
            wb_xml = z.read('xl/workbook.xml').decode('utf-8')
            rels_xml = z.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        except KeyError:
            return 0

    rels: Dict[str, str] = {}
    for m in re.finditer(r'<Relationship [^>]*/>', rels_xml):
        rid_m = re.search(r'Id="([^"]+)"', m.group(0))
        tgt_m = re.search(r'Target="([^"]+)"', m.group(0))
        if rid_m and tgt_m:
            rels[rid_m.group(1)] = tgt_m.group(1)

    sheet_to_xml: Dict[str, str] = {}
    for m in re.finditer(r'<sheet [^>]*/>', wb_xml):
        name_m = re.search(r'name="([^"]+)"', m.group(0))
        rid_m  = re.search(r'r:id="([^"]+)"', m.group(0))
        if name_m and rid_m and rid_m.group(1) in rels:
            sheet_to_xml[name_m.group(1)] = _normalize_target_path(
                rels[rid_m.group(1)]
            )

    ET.register_namespace('', _XL_NS)
    tmp_path = path.with_suffix(path.suffix + '.fix')
    total = 0
    with zipfile.ZipFile(path, 'r') as zin, \
         zipfile.ZipFile(tmp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            sheet = next(
                (s for s, x in sheet_to_xml.items() if x == item), None,
            )
            if sheet and cache.get(sheet):
                ws_cache = cache[sheet]
                try:
                    root = ET.fromstring(data)
                except ET.ParseError:
                    zout.writestr(item, data)
                    continue
                sheetdata = root.find(f'{{{_XL_NS}}}sheetData')
                if sheetdata is not None:
                    for row in sheetdata.findall(f'{{{_XL_NS}}}row'):
                        for cell in row.findall(f'{{{_XL_NS}}}c'):
                            f_el = cell.find(f'{{{_XL_NS}}}f')
                            if f_el is None:
                                continue
                            ref = cell.get('r')
                            cached = ws_cache.get(ref)
                            if cached is None:
                                continue
                            v_el = cell.find(f'{{{_XL_NS}}}v')
                            if v_el is None:
                                v_el = ET.SubElement(cell, f'{{{_XL_NS}}}v')
                            if isinstance(cached, bool):
                                v_el.text = '1' if cached else '0'
                            elif isinstance(cached, int):
                                v_el.text = str(cached)
                            else:
                                v_el.text = repr(float(cached))
                            total += 1
                    data = ET.tostring(
                        root, encoding='utf-8', xml_declaration=True,
                    )
            zout.writestr(item, data)
    shutil.move(str(tmp_path), str(path))
    return total
