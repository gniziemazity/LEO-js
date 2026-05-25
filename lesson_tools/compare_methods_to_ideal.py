"""compare_methods_to_ideal.py

Compare every diff_marks_<method>.json against diff_marks_ideal.json.

Two modes:

(1) Multi-lesson mode (default — invoked by `npm run eval`):
        python compare_methods_to_ideal.py
    A file picker asks for the Grades excel. The script then walks
    `<root>/lessons/<lesson>/` for each lesson dir, evaluates every student
    under `<lesson>/anon_ids/`, and writes a combined workbook
    `<root>/Method_Evaluation.xlsx` with:
      * Summary sheet — F1, Pair F1, Pair Mean Dist, Result, and Progress
        per method × lesson (5 columns/lesson; Result and Progress are
        filled only on the per-method Total row, since they are not
        per-label metrics)
      * one sheet per lesson — the same totals + per-student tables that
        single-project mode produces.

(2) Single-project mode (legacy):
        python compare_methods_to_ideal.py PROJECT_DIR
    PROJECT_DIR contains student subfolders directly. Teacher files are
    looked up in PROJECT_DIR itself or PROJECT_DIR/reconstructed/. Output:
    PROJECT_DIR/methods_vs_ideal.xlsx.

For each (student, method, label) where label ∈ {missing, extra, ghost_extra}:
  * mark-level TP/FP/FN/TN with precision, recall, F1, accuracy
    (universe = all non-comment tokens in the relevant source files)
  * pair-level TP/FP/FN/TN measured on TP marks: does the method's pairing or
    insert-anchor on a correctly identified mark match the ideal's? Comparison
    uses (file, start, end) of the partner — token text is not required to match.
  * pair distance: when both method and ideal made a same-kind same-file pairing
    decision, |method_target_start - ideal_target_start| in characters. The
    Pair Mean Dist column averages this over comparable pairs (exact matches
    contribute 0). Useful for distinguishing "method picked the right neighborhood
    but the wrong token" from "method was completely off."

Per (student, method) — outcome-based, independent of the label classification:
  * Result: 1 - edit_distance(method_corrected_tokens, ideal_corrected_tokens)
    / len(ideal_corrected_tokens), pooled across all student files. Each side
    is the student source with that diff_marks' actionable edits applied
    (delete extras/ghost_extras, splice missings at insert_at, replace
    paired-with partners, relocate move_to extras), tokenized with
    _CHAR_TOKEN_RE and comments stripped. Higher = method's corrections
    converge on the same end-state the ideal would produce. Orthogonal to F1:
    a method with bad pair anchoring (Pair F1 low, Pair Mean Dist high) can
    still have Result near 1.0 if its mark spans simply land on the right
    tokens. Conversely, compensating errors — an FP delete cancelled by an
    FP insert — can score artificially well here; treat as a complement to
    F1, not a replacement.
  * Progress: 1 - edit_distance(method_corrected, ideal_corrected) /
    edit_distance(student_original, ideal_corrected), pooled. Normalises the
    same numerator as Result by the *amount of work the student needed*
    rather than total file size — so unchanged-but-already-correct tokens
    don't inflate the score. Interpretation: "fraction of the student's gap
    to ideal that this method closed." 1.0 = perfect, 0.0 = method produced
    no improvement (≈ student original), negative = method made the code
    further from ideal than the student already was. Much more
    discriminating than Result on lessons where most tokens are trivially
    correct (e.g. wall).
"""

from __future__ import annotations

import difflib
import json
import sys
from collections import defaultdict
from pathlib import Path

from utils.folder_utils import CODE_EXTS, pick_file
from utils.similarity_measures import _CHAR_TOKEN_RE, _comment_ranges
from utils.token_log_mixin import (
    _LANG_EXT_LABEL,
    _effective_ext_at,
    _embedded_lang_ranges_for,
    _ext_of,
)

try:
    import pandas as pd
except ImportError:
    print("error: pandas + openpyxl required. pip install pandas openpyxl",
          file=sys.stderr)
    sys.exit(1)


IDEAL_FILE = "diff_marks_ideal.json"
LABELS = ("missing", "extra", "ghost_extra")

METHOD_LABELS = {
    "leo_star":  "LEO*",
    "leo":       "LEO",
    "lcs_star":  "LCS*",
    "lcs":       "LCS",
    "lev_star":  "Lev*",
    "lev":       "Lev",
    "ro_star":   "R/O*",
    "ro":        "R/O",
    "git_star":  "Git*",
    "git":       "Git",
}


def _load_json(path: Path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _read_text(path: Path) -> str:
    try:
        return path.read_text(encoding="utf-8", errors="ignore").replace("\r\n", "\n")
    except Exception:
        return ""


def _non_comment_token_count(text: str, ext: str | None = None) -> int:
    starts, ends = _comment_ranges(text, ext)
    spans = list(zip(starts, ends))
    n = 0
    for tm in _CHAR_TOKEN_RE.finditer(text):
        pos = tm.start()
        in_comment = False
        for lo, hi in spans:
            if lo <= pos < hi:
                in_comment = True
                break
            if pos < lo:
                break
        if not in_comment:
            n += 1
    return n


def _non_comment_tokens(text: str, ext: str | None = None) -> list[str]:
    starts, ends = _comment_ranges(text, ext)
    spans = list(zip(starts, ends))
    out: list[str] = []
    for tm in _CHAR_TOKEN_RE.finditer(text):
        pos = tm.start()
        in_comment = False
        for lo, hi in spans:
            if lo <= pos < hi:
                in_comment = True
                break
            if pos < lo:
                break
        if not in_comment:
            out.append(tm.group(0))
    return out


def _build_edit_list(student_marks_for_file, teacher_marks_flat, fname):
    """Edits to apply to student file `fname` so its content matches the post-
    correction state implied by these diff marks.

    Each edit is (start, end, replacement_text). Spans are in source code-units
    (the same offsets the marks carry)."""
    edits: list[tuple[int, int, str]] = []
    for m in student_marks_for_file or []:
        lbl = m.get("label")
        if lbl not in ("extra", "ghost_extra"):
            continue
        s = m.get("start")
        e = m.get("end")
        if s is None or e is None:
            continue
        mt = m.get("move_to")
        if isinstance(mt, dict) and mt.get("file") == fname and mt.get("pos") is not None:
            edits.append((s, e, ""))
            edits.append((mt["pos"], mt["pos"], m.get("token", "")))
        else:
            edits.append((s, e, ""))
    for m in teacher_marks_flat or []:
        if m.get("label") != "missing":
            continue
        token = m.get("token", "")
        pw = m.get("paired_with")
        ia = m.get("insert_at")
        if isinstance(pw, dict) and pw.get("file") == fname and not pw.get("ghost"):
            ps = pw.get("start")
            pe = pw.get("end")
            if ps is not None and pe is not None:
                edits.append((ps, pe, token))
                continue
        if isinstance(ia, dict) and ia.get("file") == fname and ia.get("pos") is not None:
            edits.append((ia["pos"], ia["pos"], token))
    return edits


def _apply_edits(text: str, edits) -> str:
    """Apply edits right-to-left; silently skip any whose span overlaps an
    already-applied later edit."""
    if not edits:
        return text
    ordered = sorted(edits, key=lambda e: (-e[0], -e[1]))
    out = text
    last_start = len(text) + 1
    for s, e, repl in ordered:
        if s < 0 or e < s or e > last_start:
            continue
        out = out[:s] + repl + out[e:]
        last_start = s
    return out


def _flatten_teacher_missings(data) -> list:
    out = []
    for items in (data.get("teacher_files") or {}).values():
        for m in items or []:
            if m.get("label") == "missing":
                out.append(m)
    return out


def _token_edit_distance(a: list[str], b: list[str]) -> int:
    sm = difflib.SequenceMatcher(a=a, b=b, autojunk=False)
    dist = 0
    for tag, i1, i2, j1, j2 in sm.get_opcodes():
        if tag == "equal":
            continue
        dist += max(i2 - i1, j2 - j1)
    return dist


def _find_teacher_file(project_dir: Path, fname: str) -> Path | None:
    direct = project_dir / fname
    if direct.is_file():
        return direct
    reco = project_dir / "reconstructed" / fname
    if reco.is_file():
        return reco
    start = project_dir / "start" / fname
    if start.is_file():
        return start
    correct = project_dir / "correct" / fname
    if correct.is_file():
        return correct
    return None


def _find_student_file(student_dir: Path, fname: str) -> Path | None:
    p = student_dir / fname
    return p if p.is_file() else None


def _collect_marks(data):
    out = {lbl: defaultdict(dict) for lbl in LABELS}
    for fname, items in (data.get("teacher_files") or {}).items():
        for m in items or []:
            if m.get("label") == "missing" and "start" in m and "end" in m:
                out["missing"][fname][(m["start"], m["end"])] = m
    for fname, items in (data.get("student_files") or {}).items():
        for m in items or []:
            lbl = m.get("label")
            if lbl in ("extra", "ghost_extra") and "start" in m and "end" in m:
                out[lbl][fname][(m["start"], m["end"])] = m
    return out


def _pair_signature(mark, ignore_ghost_pairs: bool = False) -> tuple | None:
    pw = mark.get("paired_with")
    if isinstance(pw, dict):
        if pw.get("ghost"):
            if ignore_ghost_pairs:
                return None
            return ("pair", pw.get("file"), pw.get("start"), pw.get("end"), "ghost")
        return ("pair", pw.get("file"), pw.get("start"), pw.get("end"))
    ia = mark.get("insert_at")
    if isinstance(ia, dict):
        return ("insert", ia.get("file"), ia.get("pos"))
    return None


def _pair_distance(ms, ts) -> int | None:
    """Char distance between method's pair target and ideal's pair target.

    Returns 0 for exact match, positive int for same-kind same-file mismatch,
    or None if incomparable (different kinds, different files, missing values)."""
    if ms is None or ts is None:
        return None
    if ms[0] != ts[0]:
        return None
    if ms[0] == "pair":
        if ms[1] != ts[1] or ms[2] is None or ts[2] is None:
            return None
        return abs(ms[2] - ts[2])
    if ms[0] == "insert":
        if ms[1] != ts[1] or ms[2] is None or ts[2] is None:
            return None
        return abs(ms[2] - ts[2])
    return None


def _is_star_method(method_key: str) -> bool:
    return method_key.endswith("_star")


def _merge_marks(*by_file_dicts):
    out: dict = defaultdict(dict)
    for d in by_file_dicts:
        for fname, marks in d.items():
            out[fname].update(marks)
    return out


def _files_referenced_by_source(named_data: dict) -> tuple[dict, dict]:
    """Return ({teacher_fname: [sources]}, {student_fname: [sources]})."""
    teacher_src: dict = {}
    student_src: dict = {}
    for source, data in named_data.items():
        if not data:
            continue
        for fname in (data.get("teacher_files") or {}):
            teacher_src.setdefault(fname, []).append(source)
        for fname in (data.get("student_files") or {}):
            student_src.setdefault(fname, []).append(source)
    return teacher_src, student_src


def _universe_size(side: str, fname: str, project_dir: Path,
                   student_dir: Path, cache: dict) -> int:
    key = (side, fname) if side == "teacher" else (side, str(student_dir), fname)
    if key in cache:
        return cache[key]
    if side == "teacher":
        path = _find_teacher_file(project_dir, fname)
    else:
        path = _find_student_file(student_dir, fname)
    if not path:
        cache[key] = 0
        return 0
    n = _non_comment_token_count(_read_text(path), path.suffix.lower())
    cache[key] = n
    return n


def _score_label_for_files(label, method_marks_by_file, ideal_marks_by_file,
                           file_universe, ignore_ghost_pairs: bool = False):
    tp = fp = fn = tn = 0
    pair_tp = pair_fp = pair_fn = pair_tn = 0
    pair_dist_sum = 0
    pair_dist_n = 0
    pair_dist_max = 0
    fnames = (
        set(method_marks_by_file) | set(ideal_marks_by_file) | set(file_universe)
    )
    for fname in fnames:
        m = method_marks_by_file.get(fname, {})
        t = ideal_marks_by_file.get(fname, {})
        m_keys = set(m)
        t_keys = set(t)
        tp_keys = m_keys & t_keys
        local_tp = len(tp_keys)
        local_fp = len(m_keys - t_keys)
        local_fn = len(t_keys - m_keys)
        univ = file_universe.get(fname, 0)
        local_tn = max(0, univ - local_tp - local_fp - local_fn)
        tp += local_tp
        fp += local_fp
        fn += local_fn
        tn += local_tn
        for k in tp_keys:
            ms = _pair_signature(m[k], ignore_ghost_pairs)
            ts = _pair_signature(t[k], ignore_ghost_pairs)
            if ms is None and ts is None:
                pair_tn += 1
            elif ms is None and ts is not None:
                pair_fn += 1
            elif ms is not None and ts is None:
                pair_fp += 1
            elif ms == ts:
                pair_tp += 1
            else:
                pair_fp += 1
                pair_fn += 1
            d = _pair_distance(ms, ts)
            if d is not None:
                pair_dist_sum += d
                pair_dist_n += 1
                if d > pair_dist_max:
                    pair_dist_max = d
    return {
        "TP": tp, "FP": fp, "FN": fn, "TN": tn,
        "Pair TP": pair_tp, "Pair FP": pair_fp,
        "Pair FN": pair_fn, "Pair TN": pair_tn,
        "Pair Dist Sum": pair_dist_sum,
        "Pair Dist N": pair_dist_n,
        "Pair Dist Max": pair_dist_max,
    }


def _add_metrics(row):
    tp, fp, fn, tn = row["TP"], row["FP"], row["FN"], row["TN"]
    p = tp / (tp + fp) if (tp + fp) else 0.0
    r = tp / (tp + fn) if (tp + fn) else 0.0
    f = 2 * p * r / (p + r) if (p + r) else 0.0
    total = tp + fp + fn + tn
    acc = (tp + tn) / total if total else 0.0
    row["Precision"] = round(p, 4)
    row["Recall"] = round(r, 4)
    row["F1"] = round(f, 4)
    row["Accuracy"] = round(acc, 4)

    ptp, pfp, pfn, ptn = row["Pair TP"], row["Pair FP"], row["Pair FN"], row["Pair TN"]
    pp = ptp / (ptp + pfp) if (ptp + pfp) else 0.0
    pr = ptp / (ptp + pfn) if (ptp + pfn) else 0.0
    pf = 2 * pp * pr / (pp + pr) if (pp + pr) else 0.0
    ptotal = ptp + pfp + pfn + ptn
    pacc = (ptp + ptn) / ptotal if ptotal else 0.0
    row["Pair Precision"] = round(pp, 4)
    row["Pair Recall"] = round(pr, 4)
    row["Pair F1"] = round(pf, 4)
    row["Pair Accuracy"] = round(pacc, 4)

    pds = row.get("Pair Dist Sum", 0)
    pdn = row.get("Pair Dist N", 0)
    row["Pair Mean Dist"] = round(pds / pdn, 2) if pdn else 0.0


def _method_display_name(key: str) -> str:
    return METHOD_LABELS.get(key, key)


def _method_sort_key(key: str) -> tuple:
    order = list(METHOD_LABELS.keys())
    try:
        return (0, order.index(key))
    except ValueError:
        return (1, key)


def _list_methods(student_dir: Path) -> list[tuple[str, Path]]:
    out = []
    for entry in sorted(student_dir.iterdir()):
        name = entry.name
        if (
            entry.is_file()
            and name.startswith("diff_marks_")
            and name.endswith(".json")
            and name != IDEAL_FILE
        ):
            method = name[len("diff_marks_") : -len(".json")]
            out.append((method, entry))
    return out


def _list_students(project_dir: Path) -> list[Path]:
    return sorted(
        d for d in project_dir.iterdir()
        if d.is_dir() and (d / IDEAL_FILE).is_file()
    )


_PER_STUDENT_COLS = [
    "Student", "Method", "Label",
    "TP", "FP", "FN", "TN",
    "Precision", "Recall", "F1", "Accuracy",
    "Pair TP", "Pair FP", "Pair FN", "Pair TN",
    "Pair Precision", "Pair Recall", "Pair F1", "Pair Accuracy",
    "Pair Dist N", "Pair Mean Dist", "Pair Dist Max",
]
_TOTALS_COLS = [
    "Method", "Label", "Students",
    "TP", "FP", "FN", "TN",
    "Precision", "Recall", "F1", "Accuracy",
    "Pair TP", "Pair FP", "Pair FN", "Pair TN",
    "Pair Precision", "Pair Recall", "Pair F1", "Pair Accuracy",
    "Pair Dist N", "Pair Mean Dist", "Pair Dist Max",
]


def evaluate(
    teacher_dir: Path, students_dir: Path | None = None,
) -> tuple[list[dict], list[dict], dict[str, dict]]:
    if students_dir is None:
        students_dir = teacher_dir
    students = _list_students(students_dir)
    if not students:
        raise SystemExit(f"No student subfolders with {IDEAL_FILE} in {students_dir}")
    project_dir = teacher_dir

    per_student: list[dict] = []
    aggregate: dict[tuple[str, str], dict] = {}
    methods_per_student_count: dict[tuple[str, str], int] = defaultdict(int)
    universe_cache: dict = {}
    result_aggregate: dict[str, dict] = {}

    for student_dir in students:
        student_id = student_dir.name
        ideal_path = student_dir / IDEAL_FILE
        ideal_data = _load_json(ideal_path)
        ideal_marks = _collect_marks(ideal_data)

        methods = _list_methods(student_dir)
        if not methods:
            print(f"  {student_id}: no diff_marks_<method>.json siblings; skipping")
            continue
        method_names = ", ".join(_method_display_name(m) for m, _ in methods)
        print(f"  {student_id}: {len(methods)} method(s) — {method_names}")

        per_method_data = {}
        for method, mpath in methods:
            try:
                per_method_data[method] = _load_json(mpath)
            except Exception as e:
                print(f"  warn: failed to load {mpath}: {e}", file=sys.stderr)

        named_data = {"ideal": ideal_data, **per_method_data}
        teacher_src, student_src = _files_referenced_by_source(named_data)

        teacher_universe = {}
        for fname in teacher_src:
            n = _universe_size(
                "teacher", fname, project_dir, student_dir, universe_cache,
            )
            if n == 0 and not _find_teacher_file(project_dir, fname):
                srcs = ", ".join(_method_display_name(s) for s in teacher_src[fname])
                print(
                    f"  warn: skipping teacher-side {fname!r} for student "
                    f"{student_id} (referenced by {srcs}; file not found)",
                    file=sys.stderr,
                )
                continue
            teacher_universe[fname] = n

        student_universe = {}
        for fname in student_src:
            n = _universe_size(
                "student", fname, project_dir, student_dir, universe_cache,
            )
            if n == 0 and not _find_student_file(student_dir, fname):
                srcs = ", ".join(_method_display_name(s) for s in student_src[fname])
                print(
                    f"  warn: skipping student-side {fname!r} for student "
                    f"{student_id} (referenced by {srcs}; file not found)",
                    file=sys.stderr,
                )
                continue
            student_universe[fname] = n

        def _filter_marks(marks_by_file, allowed):
            return {f: m for f, m in marks_by_file.items() if f in allowed}

        ideal_marks = {
            lbl: _filter_marks(
                ideal_marks[lbl],
                teacher_universe if lbl == "missing" else student_universe,
            )
            for lbl in ideal_marks
        }

        student_text_cache: dict[str, tuple[str, str]] = {}
        for fname in student_universe:
            spath = _find_student_file(student_dir, fname)
            if spath:
                student_text_cache[fname] = (
                    _read_text(spath), spath.suffix.lower(),
                )

        ideal_teacher_missings = _flatten_teacher_missings(ideal_data)
        ideal_corrected_tokens: dict[str, list[str]] = {}
        student_baseline_dist = 0
        for fname, (text, ext) in student_text_cache.items():
            ideal_student_marks = (ideal_data.get("student_files") or {}).get(fname, [])
            edits = _build_edit_list(ideal_student_marks, ideal_teacher_missings, fname)
            ideal_corrected_tokens[fname] = _non_comment_tokens(
                _apply_edits(text, edits), ext,
            )
            original_tokens = _non_comment_tokens(text, ext)
            student_baseline_dist += _token_edit_distance(
                original_tokens, ideal_corrected_tokens[fname],
            )

        for method, mdata in per_method_data.items():
            method_teacher_missings = _flatten_teacher_missings(mdata)
            student_dist = 0
            student_ideal_len = 0
            for fname, (text, ext) in student_text_cache.items():
                m_student_marks = (mdata.get("student_files") or {}).get(fname, [])
                edits = _build_edit_list(
                    m_student_marks, method_teacher_missings, fname,
                )
                method_tokens = _non_comment_tokens(_apply_edits(text, edits), ext)
                ideal_tokens = ideal_corrected_tokens.get(fname, [])
                student_dist += _token_edit_distance(method_tokens, ideal_tokens)
                student_ideal_len += len(ideal_tokens)
            ragg = result_aggregate.setdefault(method, {
                "Result Dist": 0, "Result Ideal Tokens": 0,
                "Baseline Dist": 0, "Students": 0,
            })
            ragg["Result Dist"] += student_dist
            ragg["Result Ideal Tokens"] += student_ideal_len
            ragg["Baseline Dist"] += student_baseline_dist
            ragg["Students"] += 1

            method_marks = _collect_marks(mdata)
            method_marks = {
                lbl: _filter_marks(
                    method_marks[lbl],
                    teacher_universe if lbl == "missing" else student_universe,
                )
                for lbl in method_marks
            }
            is_star = _is_star_method(method)

            label_specs: list[tuple[str, dict, dict, dict, bool]] = []
            label_specs.append((
                "missing",
                method_marks["missing"],
                ideal_marks["missing"],
                teacher_universe,
                False,
            ))
            if is_star:
                label_specs.append((
                    "extra",
                    method_marks["extra"],
                    ideal_marks["extra"],
                    student_universe,
                    False,
                ))
                label_specs.append((
                    "ghost_extra",
                    method_marks["ghost_extra"],
                    ideal_marks["ghost_extra"],
                    student_universe,
                    False,
                ))
            else:
                label_specs.append((
                    "extra (incl. ghost)",
                    _merge_marks(method_marks["extra"], method_marks["ghost_extra"]),
                    _merge_marks(ideal_marks["extra"], ideal_marks["ghost_extra"]),
                    student_universe,
                    True,
                ))

            for label, m_marks, t_marks, universe, ignore_ghost_pairs in label_specs:
                stats = _score_label_for_files(
                    label, m_marks, t_marks, universe,
                    ignore_ghost_pairs=ignore_ghost_pairs,
                )
                row = {
                    "Student": student_id,
                    "Method": _method_display_name(method),
                    "_method_key": method,
                    "Label": label,
                    **stats,
                }
                _add_metrics(row)
                per_student.append(row)

                key = (method, label)
                agg = aggregate.setdefault(key, {
                    "TP": 0, "FP": 0, "FN": 0, "TN": 0,
                    "Pair TP": 0, "Pair FP": 0, "Pair FN": 0, "Pair TN": 0,
                    "Pair Dist Sum": 0, "Pair Dist N": 0, "Pair Dist Max": 0,
                })
                for k in ("TP", "FP", "FN", "TN",
                         "Pair TP", "Pair FP", "Pair FN", "Pair TN",
                         "Pair Dist Sum", "Pair Dist N"):
                    agg[k] += stats[k]
                if stats["Pair Dist Max"] > agg["Pair Dist Max"]:
                    agg["Pair Dist Max"] = stats["Pair Dist Max"]
                methods_per_student_count[key] += 1

    totals: list[dict] = []
    for (method, label), stats in aggregate.items():
        row = {
            "Method": _method_display_name(method),
            "_method_key": method,
            "Label": label,
            "Students": methods_per_student_count[(method, label)],
            **stats,
        }
        _add_metrics(row)
        totals.append(row)

    label_order = {
        "missing": 0,
        "extra": 1,
        "extra (incl. ghost)": 1,
        "ghost_extra": 2,
    }
    per_student.sort(key=lambda r: (
        r["Student"],
        _method_sort_key(r["_method_key"]),
        label_order.get(r["Label"], 99),
    ))
    totals.sort(key=lambda r: (
        _method_sort_key(r["_method_key"]),
        label_order.get(r["Label"], 99),
    ))

    return per_student, totals, result_aggregate


def _autosize(ws, df):
    from openpyxl.utils import get_column_letter
    for col_idx, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].tolist()])
        letter = get_column_letter(col_idx)
        existing = ws.column_dimensions[letter].width or 0
        ws.column_dimensions[letter].width = max(existing, min(max(max_len + 2, 10), 32))


def write_excel(out_path: Path, per_student: list[dict], totals: list[dict]) -> None:
    ps_df = pd.DataFrame(
        [{c: r.get(c, "") for c in _PER_STUDENT_COLS} for r in per_student],
        columns=_PER_STUDENT_COLS,
    )
    tot_df = pd.DataFrame(
        [{c: r.get(c, "") for c in _TOTALS_COLS} for r in totals],
        columns=_TOTALS_COLS,
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        tot_df.to_excel(writer, sheet_name="Totals", index=False)
        ps_df.to_excel(writer, sheet_name="Per Student", index=False)
        for sheet_name, df in (("Totals", tot_df), ("Per Student", ps_df)):
            ws = writer.sheets[sheet_name]
            _autosize(ws, df)
            ws.freeze_panes = "A2"


def _aggregate_method_stats(totals: list[dict]) -> dict[str, dict]:
    agg: dict[str, dict] = {}
    for r in totals:
        key = r["_method_key"]
        a = agg.setdefault(key, {
            "Method": r["Method"],
            "TP": 0, "FP": 0, "FN": 0, "TN": 0,
            "Pair TP": 0, "Pair FP": 0, "Pair FN": 0, "Pair TN": 0,
            "Pair Dist Sum": 0, "Pair Dist N": 0, "Pair Dist Max": 0,
        })
        for k in ("TP", "FP", "FN", "TN",
                  "Pair TP", "Pair FP", "Pair FN", "Pair TN",
                  "Pair Dist Sum", "Pair Dist N"):
            a[k] += r.get(k, 0)
        max_d = r.get("Pair Dist Max", 0)
        if max_d > a["Pair Dist Max"]:
            a["Pair Dist Max"] = max_d
    return agg


def _f1(tp: int, fp: int, fn: int) -> float:
    p = tp / (tp + fp) if (tp + fp) else 0.0
    r = tp / (tp + fn) if (tp + fn) else 0.0
    return 2 * p * r / (p + r) if (p + r) else 0.0


def _safe_sheet_name(name: str, used: set[str]) -> str:
    bad = set('[]:*?/\\')
    safe = "".join(c for c in name if c not in bad)[:31] or "Sheet"
    base = safe
    i = 2
    while safe in used:
        suffix = f" ({i})"
        safe = (base[:31 - len(suffix)] + suffix)
        i += 1
    used.add(safe)
    return safe


def _evaluate_lesson_languages(lesson_dir: Path) -> dict[str, dict]:
    students_dir = lesson_dir / "anon_ids"
    if not students_dir.is_dir():
        return {}
    students = _list_students(students_dir)
    if not students:
        return {}

    valid_exts = {ext for ext, _ in _LANG_EXT_LABEL}
    teacher_tokens: dict[str, int] = {}
    missing_counts: dict[str, int] = {}
    extra_counts: dict[str, int] = {}
    ghost_counts: dict[str, int] = {}

    teacher_ranges_cache: dict[str, dict] = {}
    student_ranges_cache: dict[tuple[str, str], dict] = {}

    referenced_teacher_files: set[str] = set()
    for student_dir in students:
        ideal_data = _load_json(student_dir / IDEAL_FILE)
        for fname in (ideal_data.get("teacher_files") or {}):
            referenced_teacher_files.add(fname)

    for fname in referenced_teacher_files:
        ext = _ext_of(fname)
        if ext not in valid_exts:
            continue
        path = _find_teacher_file(lesson_dir, fname)
        if not path:
            continue
        text = _read_text(path)
        ranges = _embedded_lang_ranges_for(text, ext)
        teacher_ranges_cache[fname] = ranges
        starts, ends = _comment_ranges(text, ext)
        spans = list(zip(starts, ends))
        for tm in _CHAR_TOKEN_RE.finditer(text):
            pos = tm.start()
            in_comment = False
            for lo, hi in spans:
                if lo <= pos < hi:
                    in_comment = True
                    break
                if pos < lo:
                    break
            if in_comment:
                continue
            eff_ext = _effective_ext_at(pos, ext, ranges)
            if eff_ext in valid_exts:
                teacher_tokens[eff_ext] = teacher_tokens.get(eff_ext, 0) + 1

    for student_dir in students:
        ideal_data = _load_json(student_dir / IDEAL_FILE)
        for fname, items in (ideal_data.get("teacher_files") or {}).items():
            ext = _ext_of(fname)
            if ext not in valid_exts:
                continue
            ranges = teacher_ranges_cache.get(fname, {})
            for m in (items or []):
                if m.get("label") != "missing":
                    continue
                pos = m.get("start")
                if pos is None:
                    continue
                eff_ext = _effective_ext_at(pos, ext, ranges)
                if eff_ext in valid_exts:
                    missing_counts[eff_ext] = missing_counts.get(eff_ext, 0) + 1
        for fname, items in (ideal_data.get("student_files") or {}).items():
            ext = _ext_of(fname)
            if ext not in valid_exts:
                continue
            cache_key = (str(student_dir), fname)
            if cache_key not in student_ranges_cache:
                spath = _find_student_file(student_dir, fname)
                if spath:
                    text = _read_text(spath)
                    student_ranges_cache[cache_key] = _embedded_lang_ranges_for(text, ext)
                else:
                    student_ranges_cache[cache_key] = {}
            ranges = student_ranges_cache[cache_key]
            for m in (items or []):
                lbl = m.get("label")
                if lbl not in ("extra", "ghost_extra"):
                    continue
                pos = m.get("start")
                if pos is None:
                    continue
                eff_ext = _effective_ext_at(pos, ext, ranges)
                if eff_ext not in valid_exts:
                    continue
                if lbl == "extra":
                    extra_counts[eff_ext] = extra_counts.get(eff_ext, 0) + 1
                else:
                    ghost_counts[eff_ext] = ghost_counts.get(eff_ext, 0) + 1

    n_students = len(students)
    out: dict[str, dict] = {}
    for ext, label in _LANG_EXT_LABEL:
        tt = teacher_tokens.get(ext, 0)
        miss = missing_counts.get(ext, 0)
        ext_c = extra_counts.get(ext, 0)
        gh = ghost_counts.get(ext, 0)
        if tt == 0 and miss == 0 and ext_c == 0 and gh == 0:
            continue
        out[label] = {
            "teacher_tokens": tt,
            "students": n_students,
            "missing": miss,
            "extra": ext_c,
            "ghost": gh,
        }
    return out


_SUMMARY_LABEL_FOR_STAR = ("missing", "extra", "ghost_extra")
_SUMMARY_LABEL_FOR_PLAIN = ("missing", "extra (incl. ghost)")
_SUMMARY_LABEL_ORDER = {
    "missing": 0,
    "extra": 1,
    "extra (incl. ghost)": 1,
    "ghost_extra": 2,
    "Total": 3,
}


def _summary_metrics_from_stats(a: dict) -> tuple[float, float, float | str]:
    f1 = round(_f1(a["TP"], a["FP"], a["FN"]), 4)
    pf1 = round(_f1(a["Pair TP"], a["Pair FP"], a["Pair FN"]), 4)
    pdn = a.get("Pair Dist N", 0)
    pmd = round(a["Pair Dist Sum"] / pdn, 2) if pdn else 0.0
    return f1, pf1, pmd


def _result_score(result_agg: dict | None, mkey: str) -> float | str:
    if not result_agg:
        return ""
    r = result_agg.get(mkey)
    if not r:
        return ""
    n = r.get("Result Ideal Tokens", 0)
    if not n:
        return ""
    score = 1.0 - r.get("Result Dist", 0) / n
    if score < 0.0:
        score = 0.0
    return round(score, 4)


def _progress_score(result_agg: dict | None, mkey: str) -> float | str:
    if not result_agg:
        return ""
    r = result_agg.get(mkey)
    if not r:
        return ""
    base = r.get("Baseline Dist", 0)
    if not base:
        return ""
    return round(1.0 - r.get("Result Dist", 0) / base, 4)


def _build_summary_rows(method_keys, results):
    rows: list[dict] = []
    for _, _, mkey, mdisp in method_keys:
        labels = (_SUMMARY_LABEL_FOR_STAR if _is_star_method(mkey)
                  else _SUMMARY_LABEL_FOR_PLAIN)
        for label in labels:
            row: dict = {"Method": mdisp, "Label": label}
            for lesson_name, (_ps, tot, _ragg) in results.items():
                stats = next(
                    (r for r in tot
                     if r.get("_method_key") == mkey and r.get("Label") == label),
                    None,
                )
                if stats is None:
                    row[f"{lesson_name} F1"] = ""
                    row[f"{lesson_name} Pair F1"] = ""
                    row[f"{lesson_name} Pair Mean Dist"] = ""
                    row[f"{lesson_name} Result"] = ""
                    row[f"{lesson_name} Progress"] = ""
                    continue
                f1, pf1, pmd = _summary_metrics_from_stats(stats)
                row[f"{lesson_name} F1"] = f1
                row[f"{lesson_name} Pair F1"] = pf1
                row[f"{lesson_name} Pair Mean Dist"] = pmd
                row[f"{lesson_name} Result"] = ""
                row[f"{lesson_name} Progress"] = ""
            rows.append(row)
        row = {"Method": mdisp, "Label": "Total"}
        for lesson_name, (_ps, tot, ragg) in results.items():
            agg = _aggregate_method_stats(tot)
            a = agg.get(mkey)
            if a is None:
                row[f"{lesson_name} F1"] = ""
                row[f"{lesson_name} Pair F1"] = ""
                row[f"{lesson_name} Pair Mean Dist"] = ""
                row[f"{lesson_name} Result"] = _result_score(ragg, mkey)
                row[f"{lesson_name} Progress"] = _progress_score(ragg, mkey)
                continue
            f1, pf1, pmd = _summary_metrics_from_stats(a)
            row[f"{lesson_name} F1"] = f1
            row[f"{lesson_name} Pair F1"] = pf1
            row[f"{lesson_name} Pair Mean Dist"] = pmd
            row[f"{lesson_name} Result"] = _result_score(ragg, mkey)
            row[f"{lesson_name} Progress"] = _progress_score(ragg, mkey)
        rows.append(row)
    return rows


def _build_languages_rows(lang_stats_by_lesson: dict[str, dict]) -> list[dict]:
    rows: list[dict] = []
    for lesson_name, stats in lang_stats_by_lesson.items():
        if not stats:
            continue
        for lang, s in stats.items():
            all_marks = s["missing"] + s["extra"] + s["ghost"]
            student_tokens = s["teacher_tokens"] * s["students"]
            err = (round(all_marks / student_tokens * 100, 3)
                   if student_tokens else 0.0)
            rows.append({
                "Lesson": lesson_name,
                "Language": lang,
                "Teacher tokens": s["teacher_tokens"],
                "Students": s["students"],
                "Missing": s["missing"],
                "Extra": s["extra"],
                "Ghost": s["ghost"],
                "All marks": all_marks,
                "Errors / 100 tokens": err,
            })
        total_tokens = sum(s["teacher_tokens"] for s in stats.values())
        total_missing = sum(s["missing"] for s in stats.values())
        total_extra = sum(s["extra"] for s in stats.values())
        total_ghost = sum(s["ghost"] for s in stats.values())
        total_marks = total_missing + total_extra + total_ghost
        n_students = next(iter(stats.values()))["students"]
        student_tokens_total = total_tokens * n_students
        rows.append({
            "Lesson": lesson_name,
            "Language": "Total",
            "Teacher tokens": total_tokens,
            "Students": n_students,
            "Missing": total_missing,
            "Extra": total_extra,
            "Ghost": total_ghost,
            "All marks": total_marks,
            "Errors / 100 tokens": (
                round(total_marks / student_tokens_total * 100, 3)
                if student_tokens_total else 0.0
            ),
        })
    return rows


def write_multi_excel(
    out_path: Path,
    results: dict[str, tuple[list[dict], list[dict], dict[str, dict]]],
    lang_stats_by_lesson: dict[str, dict] | None = None,
) -> None:
    method_keys: list[tuple[int, int, str, str]] = []
    seen: set[str] = set()
    for _, (_ps, tot, _ragg) in results.items():
        for r in tot:
            key = r["_method_key"]
            if key in seen:
                continue
            seen.add(key)
            sk = _method_sort_key(key)
            method_keys.append((sk[0], sk[1] if isinstance(sk[1], int) else 999,
                                key, r["Method"]))
    method_keys.sort()

    summary_rows = _build_summary_rows(method_keys, results)
    summary_cols = ["Method", "Label"]
    for lesson_name in results:
        summary_cols.append(f"{lesson_name} F1")
        summary_cols.append(f"{lesson_name} Pair F1")
        summary_cols.append(f"{lesson_name} Pair Mean Dist")
        summary_cols.append(f"{lesson_name} Result")
        summary_cols.append(f"{lesson_name} Progress")
    summary_df = pd.DataFrame(summary_rows, columns=summary_cols)

    languages_rows = _build_languages_rows(lang_stats_by_lesson or {})
    languages_cols = ["Lesson", "Language", "Teacher tokens", "Students",
                      "Missing", "Extra", "Ghost", "All marks",
                      "Errors / 100 tokens"]
    languages_df = pd.DataFrame(languages_rows, columns=languages_cols)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        _autosize(ws, summary_df)
        ws.freeze_panes = "C2"

        if languages_rows:
            languages_df.to_excel(writer, sheet_name="Languages", index=False)
            ws = writer.sheets["Languages"]
            _autosize(ws, languages_df)
            ws.freeze_panes = "A2"

        used_sheet_names = {"Summary", "Languages"}
        for lesson_name, (per_student, totals, _result_agg) in results.items():
            sheet = _safe_sheet_name(lesson_name, used_sheet_names)
            tot_df = pd.DataFrame(
                [{c: r.get(c, "") for c in _TOTALS_COLS} for r in totals],
                columns=_TOTALS_COLS,
            )
            ps_df = pd.DataFrame(
                [{c: r.get(c, "") for c in _PER_STUDENT_COLS} for r in per_student],
                columns=_PER_STUDENT_COLS,
            )
            tot_df.to_excel(writer, sheet_name=sheet, index=False, startrow=0)
            ps_start = len(tot_df) + 3
            ps_df.to_excel(writer, sheet_name=sheet, index=False, startrow=ps_start)
            ws = writer.sheets[sheet]
            ws.cell(row=ps_start, column=1, value="Per Student").font = (
                ws.cell(row=ps_start, column=1).font.copy(bold=True)
            )
            _autosize(ws, tot_df)
            _autosize(ws, ps_df)
            ws.freeze_panes = "A2"


def _pick_grades_file() -> Path | None:
    chosen = pick_file(
        "Select the Grades Excel file",
        filetypes=[("Excel files", "*.xls *.xlsx"), ("All files", "*.*")],
    )
    return Path(chosen) if chosen else None


def _find_lessons_root(root_dir: Path) -> Path | None:
    for entry in root_dir.iterdir():
        if entry.is_dir() and entry.name.lower() == "lessons":
            return entry
    return None


def _list_lesson_dirs(lessons_root: Path) -> list[Path]:
    return sorted(d for d in lessons_root.iterdir() if d.is_dir())


def evaluate_lesson(
    lesson_dir: Path,
) -> tuple[list[dict], list[dict], dict[str, dict]]:
    students_dir = lesson_dir / "anon_ids"
    if not students_dir.is_dir():
        return [], [], {}
    return evaluate(teacher_dir=lesson_dir, students_dir=students_dir)


def run_multi(grades_path: Path) -> int:
    root = grades_path.parent
    lessons_root = _find_lessons_root(root)
    if not lessons_root:
        print(f"error: no lessons/ folder found in {root}", file=sys.stderr)
        return 1

    lesson_dirs = _list_lesson_dirs(lessons_root)
    if not lesson_dirs:
        print(f"error: no lesson subfolders in {lessons_root}", file=sys.stderr)
        return 1

    print(f"Root: {root}")
    print(f"Lessons folder: {lessons_root}")
    results: dict[str, tuple[list[dict], list[dict], dict[str, dict]]] = {}
    lang_stats: dict[str, dict] = {}
    for lesson_dir in lesson_dirs:
        lesson_name = lesson_dir.name
        print(f"\n[{lesson_name}]")
        try:
            per_student, totals, result_agg = evaluate_lesson(lesson_dir)
        except SystemExit as e:
            print(f"  skipped: {e}")
            continue
        if not totals:
            print(f"  skipped: no anon_ids/ or no methods to compare")
            continue
        results[lesson_name] = (per_student, totals, result_agg)
        lang_stats[lesson_name] = _evaluate_lesson_languages(lesson_dir)
        print(f"  {len(per_student)} per-student rows · {len(totals)} totals rows")

    if not results:
        print("\nNo lessons produced any results.", file=sys.stderr)
        return 1

    out_path = root / "Method_Evaluation.xlsx"
    write_multi_excel(out_path, results, lang_stats)
    print(f"\nWrote {out_path}")
    print(f"  {len(results)} lesson(s)")
    return 0


def main(argv: list[str]) -> int:
    if len(argv) > 1:
        project_dir = Path(argv[1]).resolve()
        if not project_dir.is_dir():
            print(f"error: {project_dir} is not a directory", file=sys.stderr)
            return 1
        print(f"Project: {project_dir}")
        per_student, totals, _result_agg = evaluate(project_dir)
        if not per_student:
            print("No methods to compare.")
            return 0
        out_path = project_dir / "methods_vs_ideal.xlsx"
        write_excel(out_path, per_student, totals)
        print(f"Wrote {out_path}")
        print(f"  {len(per_student)} per-student rows  ·  {len(totals)} totals rows")
        return 0

    grades_path = _pick_grades_file()
    if grades_path is None:
        print("No file selected.")
        return 0
    return run_multi(grades_path)


if __name__ == "__main__":
    sys.exit(main(sys.argv))
