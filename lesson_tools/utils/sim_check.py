import copy
import csv
import json
import shutil
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import load_workbook

from .similarity_measures import (
    calculate_char_histogram_similarity,
    calculate_containment,
    calculate_ide_diff_sim,
    extract_tokens,
    extract_user_identifiers,
    get_html_outside_css,
    normalize_code,
    open_csv_encoded,
    save_xlsx,
    split_code_tokens,
    split_css_tokens,
    split_follow_tokens_html,
    split_html_tokens,
)
from .token_log import TokenLogMixin
from .report_excel import ExcelReportMixin


class CodeSimilarityChecker(TokenLogMixin, ExcelReportMixin):

    def __init__(self, reference_dir: str, participation_dir: str, students_csv: str,
                 start_dir: str = None):
        self.reference_dir     = Path(reference_dir)
        self.participation_dir = Path(participation_dir)
        self.students_csv      = Path(students_csv)
        self.start_dir         = Path(start_dir) if start_dir else None

        self.student_info: Dict[str, dict] = {}
        self.name_to_id:   Dict[str, str]  = {}

        self.results:                     Dict[str, dict]              = {}
        self.student_extra_by_ext:        Dict[str, Dict[str, Counter]] = {}
        self.student_simple_extra_by_ext: Dict[str, Dict[str, Counter]] = {}
        self.simple_baseline_by_ext:      Dict[str, Counter]            = {}
        self.student_raw_texts:           Dict[str, str]                = {}

        self._student_token_stats: Dict[str, dict]    = {}
        self._student_outside_ci:  Dict[str, Counter] = {}
        self._student_all_ci:      Dict[str, Counter] = {}

        self.teacher_tokens_by_ext:           Dict[str, Counter] = {}
        self.teacher_outside_by_ext:          Dict[str, Counter] = {}
        self.teacher_inside_by_ext:           Dict[str, Counter] = {}
        self.teacher_html_outside_by_ext:     Dict[str, Counter] = {}
        self.teacher_script_outside_by_ext:   Dict[str, Counter] = {}
        self.teacher_html_outside_css_by_ext: Dict[str, Counter] = {}
        self.teacher_html_inside_css_by_ext:  Dict[str, Counter] = {}
        self._user_id_ci: set = set()

        self.remarks_data:       Dict[str, str]  = {}
        self.required_items:     List[List[str]] = []
        self.not_expected_items: List[List[str]] = []

        self._lesson_keypresses:    list = []
        self._lesson_code_inserts:  list = []
        self._lesson_interactions:  list = []
        self._lesson_all_events:    list = []
        self._lesson_session_start: int  = 0

        self._load_students()

    def _load_students(self) -> None:
        def _handle(row):
            sid = row['Student ID'].strip()
            self.student_info[sid] = {
                'name':   row['Student Name'].strip(),
                'number': row['Student Number'].strip(),
            }
            self.name_to_id[row['Student Name'].strip()] = sid

        open_csv_encoded(
            self.students_csv, _handle,
            reset_fn=lambda: (self.student_info.clear(), self.name_to_id.clear()),
        )

    def load_remarks_csv(self, csv_path: str) -> None:
        path = Path(csv_path)
        if not path.exists():
            return
        open_csv_encoded(
            path,
            lambda row: self.remarks_data.__setitem__(
                row['Student Number'].strip(), row['Remarks'].strip()
            ),
            reset_fn=self.remarks_data.clear,
        )

    def load_expected_csv(self, csv_path: str) -> None:
        path = Path(csv_path)
        if not path.exists():
            return
        self.required_items.clear()
        self.not_expected_items.clear()
        for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
            try:
                with open(path, 'r', encoding=enc, newline='') as f:
                    for row in csv.reader(f, delimiter=';'):
                        alts = [s.strip() for s in row
                                if s.strip() and not s.strip().isdigit()]
                        if not alts:
                            continue
                        if alts[0] == '!' and alts[1:]:
                            self.not_expected_items.append(alts[1:])
                        else:
                            self.required_items.append(alts)
                break
            except (UnicodeDecodeError, UnicodeError):
                self.required_items.clear()
                self.not_expected_items.clear()

    def load_required_csv(self, csv_path: str) -> None:
        self.load_expected_csv(csv_path)

    def load_lesson_json(self, project_dir: Path) -> None:
        json_files = list(project_dir.glob('*.json'))
        if len(json_files) != 1:
            if json_files:
                print(f'  Note: found {len(json_files)} JSON files in {project_dir.name}, '
                      f'skipping lesson timestamps (expected exactly 1).')
            return
        try:
            with open(json_files[0], 'r', encoding='utf-8') as f:
                data = json.load(f)
            if 'events' not in data or 'sessionStart' not in data:
                return
            evs = data['events']
            self._lesson_all_events    = evs
            self._lesson_keypresses    = [e for e in evs if 'char' in e]
            self._lesson_code_inserts  = [e for e in evs if 'code_insert' in e]
            self._lesson_interactions  = [e for e in evs if 'interaction' in e]
            self._lesson_session_start = data['sessionStart']
            _active_editor = 'main'
            for e in evs:
                if 'switch_editor' in e:
                    _active_editor = e['switch_editor']
                elif 'char' in e or 'code_insert' in e or 'anchor' in e:
                    e['editor'] = _active_editor
            print(f'  Loaded lesson log: {json_files[0].name} '
                  f'({len(self._lesson_keypresses)} keypresses, '
                  f'{len(self._lesson_code_inserts)} code inserts)')
        except Exception as e:
            print(f'  Warning: could not load lesson JSON: {e}')


    def get_code_files(self, directory: Path) -> Dict[str, Path]:
        files = {}
        for ext in ('.html', '.css', '.js'):
            matching = list(directory.glob(f'*{ext}'))
            if matching:
                files[ext] = matching[0]
        return files

    def compare_files(self,
                      ref_file: Path, student_file: Path,
                      teacher_tokens: Counter,
                      baseline_outside: Counter,
                      baseline_inside: Counter,
                      ext: str = '',
                      teacher_html_outside: Counter = None,
                      teacher_script_outside: Counter = None,
                      teacher_inside: Counter = None,
                      baseline_html_outside: Counter = None,
                      baseline_script_outside: Counter = None,
                      baseline_html_outside_css: Counter = None,
                      teacher_html_outside_css: Counter = None) -> Dict:
        try:
            ref_raw     = ref_file.read_text(encoding='utf-8', errors='ignore')
            student_raw = student_file.read_text(encoding='utf-8', errors='ignore')
            ref_lines   = normalize_code(ref_raw)
            stu_lines   = normalize_code(student_raw)

            ide_sim        = calculate_ide_diff_sim(ref_lines, stu_lines)
            char_hist_sim  = calculate_char_histogram_similarity(ref_lines, stu_lines)
            student_tokens = extract_tokens(stu_lines)

            s_html_out = s_script_out = s_html_out_css = Counter()
            if ext == '.html':
                s_html_out, s_script_out, student_inside = split_html_tokens(student_raw)
                student_outside = s_html_out + s_script_out
                s_html_out_css  = get_html_outside_css(student_raw)
            elif ext == '.css':
                student_outside, student_inside = split_css_tokens(student_raw)
            else:
                student_outside, student_inside = split_code_tokens(student_raw)

            if ext == '.html' and teacher_html_outside is not None:
                t_inc   = ((teacher_html_outside_css or teacher_html_outside)
                           + teacher_script_outside)
                s_inc   = s_html_out_css + s_script_out
                inc_sim = calculate_containment(t_inc, s_inc)
            elif ext == '.css':
                inc_sim = calculate_containment(teacher_tokens, student_outside)
            else:
                inc_sim = calculate_containment(teacher_tokens, student_outside)

            if ext == '.html' and baseline_html_outside is not None:
                bl_html       = baseline_html_outside_css or baseline_html_outside
                extra_outside = (s_html_out_css - bl_html
                                 + s_script_out - baseline_script_outside)
                extra_inside  = student_inside - baseline_inside
            elif ext == '.css':
                extra_outside = student_outside - baseline_outside
                extra_inside  = student_inside  - baseline_inside
            else:
                extra_outside = student_outside - baseline_outside
                extra_inside  = student_inside  - baseline_inside

            total = sum(student_tokens.values())
            def _pct(c): return round(sum(c.values()) / total * 100, 1) if total else 0.0
            def _fmt(c): return [f'{kw} (x{n})' if n > 1 else kw
                                 for kw, n in sorted(c.items())]
            return {
                'status':                   'success',
                'file_name':                student_file.name,
                'ide_sim':                  round(ide_sim * 100, 1),
                'char_hist_sim':            round(char_hist_sim * 100, 1),
                'inc_sim':                  inc_sim,
                'extra_outside_pct':        _pct(extra_outside),
                'extra_inside_pct':         _pct(extra_inside),
                'extra_outside_keywords':   _fmt(extra_outside),
                'extra_inside_keywords':    _fmt(extra_inside),
                'extra_outside':            extra_outside,
                'extra_inside':             extra_inside,
                'extra_counter':            extra_outside + extra_inside,
                'student_outside':          student_outside,
                'student_inside':           student_inside,
                'student_html_outside':     s_html_out,
                'student_html_outside_css': s_html_out_css,
                'student_script_outside':   s_script_out,
                'student_tokens':           student_tokens,
            }
        except Exception:
            return {'status': 'error'}

    def run_check(self) -> None:
        ref_files = self.get_code_files(self.reference_dir)

        t_outside:      Dict[str, Counter] = {}
        t_inside:       Dict[str, Counter] = {}
        t_tokens:       Dict[str, Counter] = {}
        t_html_out:     Dict[str, Counter] = {}
        t_script_out:   Dict[str, Counter] = {}
        t_html_css_out: Dict[str, Counter] = {}
        t_html_css_ins: Dict[str, Counter] = {}

        for ext, ref_file in ref_files.items():
            raw = ref_file.read_text(encoding='utf-8', errors='ignore')
            if ext == '.html':
                html_out, script_out, ins = split_html_tokens(raw)
                out = html_out + script_out
                t_html_out[ext]     = html_out
                t_script_out[ext]   = script_out
                t_html_css_out[ext] = get_html_outside_css(raw)
                _, css_ins = split_follow_tokens_html(raw)
                t_html_css_ins[ext] = css_ins
            elif ext == '.css':
                out, ins = split_css_tokens(raw)
            else:
                out, ins = split_code_tokens(raw)
            t_outside[ext] = out
            t_inside[ext]  = ins
            t_tokens[ext]  = out + ins

        self.teacher_tokens_by_ext           = t_tokens
        self.teacher_outside_by_ext          = t_outside
        self.teacher_inside_by_ext           = t_inside
        self.teacher_html_outside_by_ext     = t_html_out
        self.teacher_script_outside_by_ext   = t_script_out
        self.teacher_html_outside_css_by_ext = t_html_css_out
        self.teacher_html_inside_css_by_ext  = t_html_css_ins

        self._user_id_ci = {
            uid
            for ext, f in ref_files.items()
            for uid in extract_user_identifiers(
                f.read_text(encoding='utf-8', errors='ignore'), ext
            )
        }

        (bl_outside, bl_inside,
         bl_html_out, bl_script_out, bl_html_css) = self._build_baselines(
            ref_files, t_outside, t_inside, t_html_out, t_script_out, t_html_css_out
        )

        for student_dir in sorted(d for d in self.participation_dir.iterdir() if d.is_dir()):
            sid = self.name_to_id.get(student_dir.name)
            if sid is None:
                continue

            res        = {'files_compared': {}}
            ext_extras: Dict[str, Counter] = {}
            stu_files  = self.get_code_files(student_dir)
            html_ref   = ref_files.get('.html')

            for ext, ref_file in ref_files.items():
                if ext in stu_files:
                    result = self.compare_files(
                        ref_file, stu_files[ext],
                        t_outside.get(ext, Counter()),
                        bl_outside.get(ext, t_outside.get(ext, Counter())),
                        bl_inside.get(ext, t_inside.get(ext, Counter())),
                        ext=ext,
                        teacher_html_outside=t_html_out.get(ext),
                        teacher_script_outside=t_script_out.get(ext),
                        teacher_inside=t_inside.get(ext),
                        baseline_html_outside=bl_html_out.get(ext, t_html_out.get(ext, Counter())),
                        baseline_script_outside=bl_script_out.get(ext, t_script_out.get(ext, Counter())),
                        baseline_html_outside_css=bl_html_css.get(ext, t_html_css_out.get(ext)),
                        teacher_html_outside_css=t_html_css_out.get(ext),
                    )
                    res['files_compared'][ext] = result
                    ext_extras[ext] = result.get('extra_outside', Counter())
                else:
                    ext_extras[ext] = Counter()

            for _ext in ({'.css', '.js'} - set(ref_files.keys())):
                if _ext not in stu_files:
                    continue
                s_file = stu_files[_ext]
                if html_ref is not None and _ext == '.js':
                    t_script = t_script_out.get('.html', Counter())
                    if t_script:
                        s_raw   = s_file.read_text(encoding='utf-8', errors='ignore')
                        s_out, s_ins = split_code_tokens(s_raw)
                        _bl_out = bl_script_out.get('.html', t_script_out.get('.html', Counter()))
                        _bl_ins = bl_inside.get('.html', t_inside.get('.html', Counter()))
                        extra_out = s_out - _bl_out
                        extra_ins = s_ins - _bl_ins
                        total     = max(1, sum(extract_tokens(normalize_code(s_raw)).values()))
                        res['files_compared'][_ext] = self._no_ref_result(
                            s_file, s_out, s_ins, extra_out, extra_ins,
                            calculate_containment(t_script, s_out), total, is_script=True)
                        ext_extras[_ext] = extra_out
                        continue
                elif html_ref is not None and _ext == '.css':
                    t_html_ref = t_html_out.get('.html', Counter())
                    if t_html_ref:
                        s_raw = s_file.read_text(encoding='utf-8', errors='ignore')
                        s_out, s_ins = split_css_tokens(s_raw)
                        _bl_html = bl_html_out.get('.html', t_html_out.get('.html', Counter()))
                        _bl_ins  = bl_inside.get('.html', t_inside.get('.html', Counter()))
                        extra_out = s_out - _bl_html
                        extra_ins = s_ins - _bl_ins
                        total     = max(1, sum(extract_tokens(normalize_code(s_raw)).values()))
                        res['files_compared'][_ext] = self._no_ref_result(
                            s_file, s_out, s_ins, extra_out, extra_ins,
                            calculate_containment(t_html_ref, s_out),
                            total, is_script=False)
                        ext_extras[_ext] = extra_out
                        continue
                res['files_compared'][_ext] = {'status': 'no_reference',
                                               'file_name': s_file.name}

            self.results[sid]              = res
            self.student_extra_by_ext[sid] = ext_extras

            raw_parts:      List[str]          = []
            simple_extras:  Dict[str, Counter] = {}
            stu_outside_ci: Counter            = Counter()
            stu_all_ci:     Counter            = Counter()

            for ext in ['.html', '.css', '.js']:
                _files = list(student_dir.glob(f'*{ext}'))
                if not _files:
                    simple_extras[ext] = Counter()
                    continue
                try:
                    raw = _files[0].read_text(encoding='utf-8', errors='ignore')
                    raw_parts.append(raw)
                    if ext == '.html':
                        s_out_css    = get_html_outside_css(raw)
                        _, sc, _     = split_html_tokens(raw)
                        _, ins_css   = split_follow_tokens_html(raw)
                        stu_outside_ci += s_out_css + sc
                        stu_all_ci     += s_out_css + sc + ins_css
                        s_out = s_out_css
                    elif ext == '.css':
                        s_out, s_ins = split_css_tokens(raw)
                        stu_outside_ci += s_out
                        stu_all_ci     += s_out + s_ins
                    else:
                        s_out, s_ins = split_code_tokens(raw)
                        stu_outside_ci += s_out
                        stu_all_ci     += s_out + s_ins
                    simple_extras[ext] = (
                        s_out
                        - self.simple_baseline_by_ext.get(ext, Counter())
                    )
                except Exception:
                    simple_extras[ext] = Counter()

            self.student_simple_extra_by_ext[sid] = simple_extras
            self.student_raw_texts[sid]           = ', '.join(raw_parts)
            self._student_outside_ci[sid]         = stu_outside_ci
            self._student_all_ci[sid]             = stu_all_ci

    def _build_baselines(self, ref_files, t_outside, t_inside,
                         t_html_outside, t_script_outside, t_html_css_outside):
        bl_outside:  Dict[str, Counter] = {}
        bl_inside:   Dict[str, Counter] = {}
        bl_html_out: Dict[str, Counter] = {}
        bl_script:   Dict[str, Counter] = {}
        bl_html_css: Dict[str, Counter] = {}

        if self.start_dir and self.start_dir.is_dir():
            for ext, sf in self.get_code_files(self.start_dir).items():
                raw = sf.read_text(encoding='utf-8', errors='ignore')
                if ext == '.html':
                    h_out, sc_out, ins = split_html_tokens(raw)
                    bl_outside[ext]  = h_out + sc_out
                    bl_inside[ext]   = ins
                    bl_html_out[ext] = h_out
                    bl_script[ext]   = sc_out
                    bl_html_css[ext] = get_html_outside_css(raw)
                elif ext == '.css':
                    bl_outside[ext], bl_inside[ext] = split_css_tokens(raw)
                else:
                    bl_outside[ext], bl_inside[ext] = split_code_tokens(raw)

            def _floor(bl_dict, ref_dict):
                for ext in list(bl_dict.keys()):
                    for tok, cnt in ref_dict.get(ext, Counter()).items():
                        if cnt > bl_dict[ext].get(tok, 0):
                            bl_dict[ext][tok] = cnt

            _floor(bl_outside, t_outside)
            _floor(bl_inside,  t_inside)
            _floor(bl_html_out, t_html_outside)
            _floor(bl_script,   t_script_outside)
            _floor(bl_html_css, t_html_css_outside)

        for ext in ['.html', '.css', '.js']:
            src = None
            if self.start_dir and self.start_dir.is_dir():
                sf_list = list(self.start_dir.glob(f'*{ext}'))
                if sf_list:
                    src = sf_list[0]
            if src is None:
                src = ref_files.get(ext)
            if src:
                raw = src.read_text(encoding='utf-8', errors='ignore')
                bl  = (get_html_outside_css(raw)       if ext == '.html'
                       else split_css_tokens(raw)[0]   if ext == '.css'
                       else split_code_tokens(raw)[0])
                self.simple_baseline_by_ext[ext] = Counter(bl)
            else:
                self.simple_baseline_by_ext[ext] = Counter()

            ref_file = ref_files.get(ext)
            if ref_file:
                raw     = ref_file.read_text(encoding='utf-8', errors='ignore')
                ref_out = (get_html_outside_css(raw)       if ext == '.html'
                           else split_css_tokens(raw)[0]   if ext == '.css'
                           else split_code_tokens(raw)[0])
                for tok, cnt in ref_out.items():
                    if cnt > self.simple_baseline_by_ext[ext].get(tok, 0):
                        self.simple_baseline_by_ext[ext][tok] = cnt

        return bl_outside, bl_inside, bl_html_out, bl_script, bl_html_css

    @staticmethod
    def _no_ref_result(s_file: Path, s_out: Counter, s_ins: Counter,
                       extra_out: Counter, extra_ins: Counter,
                       inc_sim: float, total: int, is_script: bool) -> Dict:
        def _pct(c): return round(sum(c.values()) / total * 100, 1)
        def _fmt(c): return [f'{k} (x{n})' if n > 1 else k for k, n in sorted(c.items())]
        return {
            'status':                 'success',
            'file_name':              s_file.name,
            'ide_sim':                0.0,
            'char_hist_sim':          0.0,
            'inc_sim':                inc_sim,
            'extra_outside_pct':      _pct(extra_out),
            'extra_inside_pct':       _pct(extra_ins),
            'extra_outside_keywords': _fmt(extra_out),
            'extra_inside_keywords':  _fmt(extra_ins),
            'extra_outside':          extra_out,
            'extra_inside':           extra_ins,
            'student_outside':        s_out,
            'student_inside':         s_ins,
            'student_html_outside':   Counter(),
            'student_script_outside': s_out if is_script else Counter(),
            'student_tokens':         Counter(),
        }

def main() -> None:
    if len(sys.argv) < 2:
        print('Usage: sim_check.py <project_dir>')
        sys.exit(1)

    current_dir    = Path(sys.argv[1]).resolve()
    scripts_dir    = Path(__file__).resolve().parent
    correct_dir    = current_dir / 'correct'
    anon_names_dir = current_dir / 'anon_names'
    names_dir      = current_dir / 'students'
    students_csv   = current_dir.parent.parent / 'students.csv'
    start_dir      = current_dir / 'start'

    missing = [p for p in (correct_dir, anon_names_dir, students_csv) if not p.exists()]
    if missing:
        print(f'Missing: {", ".join(str(p) for p in missing)}')
        return

    checker = CodeSimilarityChecker(
        str(correct_dir), str(anon_names_dir), str(students_csv),
        start_dir=str(start_dir) if start_dir.exists() else None,
    )
    checker.run_check()

    remarks_csv = current_dir / 'remarks.csv'
    checker.load_remarks_csv(str(remarks_csv))
    if remarks_csv.exists():
        remarks_csv.unlink()

    expected_csv = current_dir / 'expected.csv'
    if not expected_csv.exists():
        expected_csv = current_dir / 'required.csv'
    checker.load_expected_csv(str(expected_csv))
    checker.load_lesson_json(current_dir)

    folder_name = current_dir.name
    print('\nWriting diff mark files...')
    if checker._lesson_keypresses:
        checker.write_keyword_log()
        checker.write_student_token_files(names_dir, anon_names_dir)
    else:
        checker.write_similarity_diff_marks(names_dir, anon_names_dir)

    checker.generate_excel_report(
        str(current_dir / f'teacher_similarity_{folder_name}.xlsx')
    )
    remarks_path = current_dir / f'remarks_{folder_name}.xlsx'
    checker.generate_remarks_report(str(remarks_path))

    grades_ts   = int(datetime.now().timestamp())
    grades_path = current_dir / f'grades_{folder_name}_{grades_ts}.xlsx'
    shutil.copy2(str(remarks_path), str(grades_path))
    _merge_existing_grades(current_dir, folder_name, grades_path)

    print(f'Done — teacher_similarity_{folder_name}.xlsx, '
          f'remarks_{folder_name}.xlsx and '
          f'grades_{folder_name}_{grades_ts}.xlsx generated.')


def _merge_existing_grades(current_dir: Path, folder_name: str, grades_path: Path) -> None:
    existing = sorted(
        (p for p in current_dir.glob(f'grades_{folder_name}_*.xlsx') if p != grades_path),
        key=lambda p: p.stat().st_mtime, reverse=True,
    )
    if not existing:
        return
    try:
        src_wb = load_workbook(existing[0])
        dst_wb = load_workbook(grades_path)
        sheet  = 'Remarks'
        if sheet not in src_wb.sheetnames or sheet not in dst_wb.sheetnames:
            return
        src_ws, dst_ws = src_wb[sheet], dst_wb[sheet]

        def _hdr(ws, *names):
            return {c.value: c.column for c in ws[1] if c.value in names}

        src_c = _hdr(src_ws, 'Obs', 'Grade', 'Comments')
        dst_c = _hdr(dst_ws, 'Obs', 'Grade', 'Comments')
        if not (src_c and dst_c):
            return

        src_map = {
            str(row[0].value): {n: row[col - 1] for n, col in src_c.items()}
            for row in src_ws.iter_rows(min_row=2)
            if row[0].value is not None
        }
        for row in dst_ws.iter_rows(min_row=2):
            sid = str(row[0].value) if row[0].value is not None else None
            if sid not in src_map:
                continue
            for name, dst_col in dst_c.items():
                sc = src_map[sid].get(name)
                if sc is None:
                    continue
                dc = row[dst_col - 1]
                dc.value = sc.value
                if sc.has_style:
                    import copy
                    dc.font          = copy.copy(sc.font)
                    dc.fill          = copy.copy(sc.fill)
                    dc.border        = copy.copy(sc.border)
                    dc.alignment     = copy.copy(sc.alignment)
                    dc.number_format = sc.number_format

        save_xlsx(dst_wb, str(grades_path), vml_source=str(grades_path))
        print(f'  Merged grades from: {existing[0].name}')
        existing[0].unlink()
        print(f'  Deleted: {existing[0].name}')
    except Exception as e:
        print(f'  Warning: could not merge existing grades: {e}')


if __name__ == '__main__':
    main()