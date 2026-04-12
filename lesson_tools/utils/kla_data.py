import json
import re
from datetime import datetime

from .kla_config import Config


def normalize_answered_by(value):
    if not value:
        return []
    parts = [p.strip() for p in str(value).split(',') if p.strip()]
    return parts


def load_keypress_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data


def extract_interactions(key_presses):
    interactions = {
        'teacher-question': [],
        'student-question': [],
        'providing-help': []
    }
    for kp in key_presses:
        if 'interaction' in kp:
            itype = kp['interaction']
            if itype in interactions:
                entry = {
                    'timestamp': kp['timestamp'] / 1000,
                    'info': kp.get('info', ''),
                    'asked_by': kp.get('asked_by', ''),
                    'answered_by': normalize_answered_by(kp.get('answered_by', None)),
                    'student': kp.get('student', ''),
                    'closed_at': kp.get('closed_at', None),
                }
                interactions[itype].append(entry)
    return interactions


def calculate_intervals(key_presses):
    intervals = []
    for i in range(1, len(key_presses)):
        intervals.append((key_presses[i]['timestamp'] - key_presses[i-1]['timestamp']) / 1000)
    return intervals


def analyze_typing_blocks(key_presses):
    if not key_presses:
        return [], [], [], []

    char_presses = [kp for kp in key_presses if 'char' in kp]
    if not char_presses:
        return [], [], [], []

    blocks = []
    current_block = [char_presses[0]]

    for i in range(1, len(char_presses)):
        time_gap = (char_presses[i]['timestamp'] - char_presses[i-1]['timestamp']) / 1000

        if time_gap < Config.BURST_GAP_THRESHOLD:
            current_block.append(char_presses[i])
        else:
            if len(current_block) >= Config.MIN_BURST_CHARS:
                blocks.append(current_block)
            current_block = [char_presses[i]]

    if len(current_block) >= Config.MIN_BURST_CHARS:
        blocks.append(current_block)

    timestamps = []
    rates = []
    texts = []
    widths = []

    _DELETE_CHARS_SET = {'\u21a2', '\u21a3', '\u26d4', '\u232b', '\u2326'}

    for block in blocks:
        start_time = block[0]['timestamp'] / 1000
        end_time = block[-1]['timestamp'] / 1000
        duration = end_time - start_time

        if duration == 0:
            duration = 1

        char_count = len(block)
        rate = (char_count / duration) * 60
        text = "".join([kp.get('char', '') for kp in block])

        center_time = (start_time + end_time) / 2
        width_days = duration / (24 * 3600)

        has_dev    = any(kp.get('_editor') == 'dev' for kp in block)
        has_remove = any(kp.get('char', '') in _DELETE_CHARS_SET for kp in block)

        timestamps.append(center_time)
        rates.append(rate)
        texts.append(text)
        widths.append(width_days)

    return timestamps, rates, texts, widths


def _parse_all_events_from_follow(follow_text, session_date):
    events = []
    for m in re.finditer(r'([^\(,]+?)\s*\((\d{2}:\d{2}:\d{2})\)', follow_text):
        label = m.group(1).strip()
        try:
            t = datetime.strptime(m.group(2), '%H:%M:%S').time()
            events.append((label, datetime.combine(session_date, t)))
        except ValueError:
            pass
    return events


def load_student_data_from_xlsx(remarks_path, similarity_path, session_start_ts, session_end_ts=None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('openpyxl not available – cannot load student xlsx data.')
        return []

    session_dt   = datetime.fromtimestamp(session_start_ts)
    session_date = session_dt.date()

    follow_data = {}
    try:
        wb_r = load_workbook(remarks_path, read_only=True, data_only=True)
        ws_r = wb_r['Remarks']
        rows_r = list(ws_r.iter_rows(values_only=True))
        wb_r.close()
        header_r      = list(rows_r[0]) if rows_r else []
        name_col_r    = None
        follow_num_col  = None
        follow_text_col = None
        for ci, h in enumerate(header_r):
            if h == 'Student' and name_col_r is None:
                name_col_r = ci
            elif h == 'Follow (E)' and follow_num_col is None:
                follow_num_col = ci
            elif h == 'Follow (E) Desc' and follow_text_col is None:
                follow_text_col = ci
        if name_col_r is None or follow_text_col is None:
            print('  Warning: could not locate Student / Follow (E) Desc columns in remarks.xlsx')
        else:
            for row in rows_r[1:]:
                name = str(row[name_col_r]).strip() if row[name_col_r] else ''
                ft   = row[follow_text_col]
                ft_str = str(ft).strip() if ft else ''
                fn   = row[follow_num_col] if follow_num_col is not None else None
                try:
                    fn_val = float(fn) if fn is not None and fn != '' else None
                except (TypeError, ValueError):
                    fn_val = None
                if name:
                    follow_data[name] = {
                        'events': (_parse_all_events_from_follow(ft_str, session_date) if ft_str else []),
                        'pct': fn_val,
                    }
    except Exception as e:
        print(f'  Error reading remarks.xlsx: {e}')

    inc_data = {}
    try:
        wb_s = load_workbook(similarity_path, read_only=True, data_only=True)
        ws_s = wb_s.active
        rows_s = list(ws_s.iter_rows(values_only=True))
        wb_s.close()
        header_s   = list(rows_s[0]) if rows_s else []
        name_col_s = None
        inc_cols   = []
        for ci, h in enumerate(header_s):
            if h == 'Student' and name_col_s is None:
                name_col_s = ci
            elif h == 'Inc' or (h and str(h).endswith('_Inc')):
                inc_cols.append(ci)
        if name_col_s is None or not inc_cols:
            print('  Warning: could not locate Student / Inc columns in teacher_similarity.xlsx')
        else:
            for row in rows_s[1:]:
                name = str(row[name_col_s]).strip() if row[name_col_s] else ''
                vals = [float(row[ci]) for ci in inc_cols
                        if row[ci] is not None and row[ci] != '']
                if name and vals:
                    inc_data[name] = sum(vals) / len(vals)
    except Exception as e:
        print(f'  Error reading teacher_similarity.xlsx: {e}')

    students = []
    all_names = set(follow_data.keys()) | set(inc_data.keys())
    session_end_dt = datetime.fromtimestamp(session_end_ts) if session_end_ts else None
    for name in sorted(all_names):
        fd = follow_data.get(name, {})
        follow_pct = fd.get('pct') if isinstance(fd, dict) else None
        if follow_pct is None:
            continue
        events = fd.get('events', []) if isinstance(fd, dict) else fd
        inc = inc_data.get(name)
        if not events:
            if session_end_dt is None:
                continue
            events = [('(followed till end)', session_end_dt)]
        follow_dt = events[0][1] if events else None
        students.append({
            'name': name,
            'follow_dt': follow_dt,
            'follow_events': events,
            'inc_sim': inc,
            'follow_pct': follow_pct,
        })
    return students
