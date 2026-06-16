import sys
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from .similarity_measures import (
    normalize_code,
    calculate_ide_diff_sim, calculate_char_histogram_similarity,
    split_code_tokens, calculate_containment,
    save_xlsx,
)
from .lesson_log import load_lesson_log
from .lv_editor import reconstruct_all_with_ghosts
from .folder_utils import LANG_EXTS


_HEADER_ROW_HEIGHT = 50
_BLACK_BORDER = Border(
    left=Side(style='medium', color='000000'), right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),  bottom=Side(style='medium', color='000000'),
)


def _parse_tokens_txt(path: Path) -> Tuple[Counter, Counter]:
    extra_outside: Counter = Counter()
    extra_comment: Counter = Counter()
    try:
        for line in path.read_text(encoding='utf-8', errors='ignore').splitlines():
            if line.startswith('#') or not line.strip():
                continue
            parts = line.split('\t')
            if len(parts) < 2:
                continue
            token = parts[0]
            flags = set(parts[2:]) if len(parts) > 2 else set()
            if 'EXTRA' in flags:
                if 'COMMENT' in flags:
                    extra_comment[token] += 1
                else:
                    extra_outside[token] += 1
    except Exception:
        pass
    return extra_outside, extra_comment





def _extra_comment(cA: Counter, cB: Counter, name_a: str, fmt=str) -> Optional[Comment]:
    inter = cA & cB
    if not inter:
        return None
    lines = [f'{fmt(kw)} (x{n})' if n > 1 else fmt(kw) for kw, n in sorted(inter.items())]
    only_a = cA - cB
    if only_a:
        lines.append(f'\n-- {name_a} ONLY --\n')
        lines.extend(f'{fmt(kw)} (x{n})' if n > 1 else fmt(kw) for kw, n in sorted(only_a.items()))
    cm = Comment(', '.join(lines).replace('\n,', '\n'), 'peer_sim')
    cm.width  = 500
    cm.height = min(100 + 30 * len(lines), 4000)
    return cm


class PeerSimilarityChecker:
    def __init__(self, students_dir: str, teacher_dir: str,
                 start_dir: str = None,
                 id_map: Optional[Dict[str, str]] = None,
                 events: Optional[list] = None,
                 lesson_file: Optional[str] = None):
        self.students_dir = Path(students_dir)
        self.teacher_dir  = Path(teacher_dir)
        self.start_dir    = Path(start_dir) if start_dir else None
        self.id_map       = id_map or {}
        self.events       = events
        self.lesson_file  = lesson_file
        self.student_data:           Dict[str, Dict[str, Optional[List[str]]]] = {}
        self.student_extra_outside:  Dict[str, Dict[str, Counter]] = {}
        self.student_extra_inside:   Dict[str, Dict[str, Counter]] = {}
        self.student_outside_full:   Dict[str, Counter] = {}
        self.baseline_outside: Dict[str, Counter] = {}
        self.baseline_inside:  Dict[str, Counter] = {}
        self.extensions = list(LANG_EXTS)

    def _read_file(self, directory: Path, ext: str) -> Optional[str]:
        files = list(directory.glob(f'*{ext}'))
        return files[0].read_text(encoding='utf-8', errors='ignore') if files else None

    def _load_baseline(self) -> None:
        for ext in self.extensions:
            src_raw = (self._read_file(self.start_dir, ext)
                       if self.start_dir and self.start_dir.is_dir() else None)
            if src_raw is None:
                src_raw = self._read_file(self.teacher_dir, ext)

            if src_raw is not None:
                self.baseline_outside[ext], self.baseline_inside[ext] = \
                    split_code_tokens(src_raw)
            else:
                self.baseline_outside[ext] = Counter()
                self.baseline_inside[ext]  = Counter()

            t_raw = self._read_file(self.teacher_dir, ext)
            if t_raw is not None:
                t_out, t_ins = split_code_tokens(t_raw)
                for tok, cnt in t_out.items():
                    if cnt > self.baseline_outside[ext].get(tok, 0):
                        self.baseline_outside[ext][tok] = cnt
                for tok, cnt in t_ins.items():
                    if cnt > self.baseline_inside[ext].get(tok, 0):
                        self.baseline_inside[ext][tok] = cnt

        if self.events:
            self._augment_baseline_with_ghosts()

    def _augment_baseline_with_ghosts(self) -> None:
        try:
            reco = reconstruct_all_with_ghosts(
                self.events, lesson_file=self.lesson_file,
            )
        except Exception:
            return
        for tab_key, info in reco.items():
            ghosts = info.get('ghosts', []) if isinstance(info, dict) else []
            for g in ghosts:
                gtext = g.get('text', '') if isinstance(g, dict) else ''
                if not gtext.strip():
                    continue
                g_out, g_ins = split_code_tokens(gtext)
                for ext in self.extensions:
                    for tok, cnt in g_out.items():
                        if cnt > self.baseline_outside[ext].get(tok, 0):
                            self.baseline_outside[ext][tok] = cnt
                    for tok, cnt in g_ins.items():
                        if cnt > self.baseline_inside[ext].get(tok, 0):
                            self.baseline_inside[ext][tok] = cnt

    def load_student_data(self) -> None:
        print("Loading student files...")
        self._load_baseline()

        for s_dir in sorted(d for d in self.students_dir.iterdir() if d.is_dir()):
            name = s_dir.name
            self.student_data[name]           = {}
            self.student_extra_outside[name]  = {}
            self.student_extra_inside[name]   = {}

            tokens_path = s_dir / 'tokens.txt'
            if tokens_path.exists():
                tok_extra_out, tok_extra_in = _parse_tokens_txt(tokens_path)
                self.student_extra_outside[name] = {'_tokens': tok_extra_out}
                self.student_extra_inside[name]  = {'_tokens': tok_extra_in}
                full_outside: Counter = Counter()
                for ext in self.extensions:
                    raw = self._read_file(s_dir, ext)
                    self.student_data[name][ext] = normalize_code(raw) if raw else None
                    if raw:
                        out, _ = split_code_tokens(raw)
                        full_outside += out
                self.student_outside_full[name] = full_outside
                continue

            full_outside = Counter()
            for ext in self.extensions:
                raw = self._read_file(s_dir, ext)
                if raw is None:
                    self.student_data[name][ext]          = None
                    self.student_extra_outside[name][ext] = Counter()
                    self.student_extra_inside[name][ext]  = Counter()
                    continue
                try:
                    self.student_data[name][ext] = normalize_code(raw)
                    out, ins = split_code_tokens(raw)
                    self.student_extra_outside[name][ext] = out - self.baseline_outside[ext]
                    self.student_extra_inside[name][ext] = ins - self.baseline_inside[ext]
                    full_outside += out
                except Exception:
                    self.student_data[name][ext]          = None
                    self.student_extra_outside[name][ext] = Counter()
                    self.student_extra_inside[name][ext]  = Counter()
            self.student_outside_full[name] = full_outside

    def _fill_matrix_sheet(
        self,
        wb: Workbook,
        title: str,
        student_names: List[str],
        scorer: Callable[[str, str], Tuple[Optional[float], Optional[Comment]]],
    ) -> None:
        ws = wb.create_sheet(title=title)
        has_ids = bool(self.id_map)
        student_ids = ([self.id_map.get(n, '') for n in student_names]
                       if has_ids else None)
        self._init_matrix_header(ws, student_names, student_ids)
        row_off = 3 if has_ids else 2
        col_off = 3 if has_ids else 2
        for r, sA in enumerate(student_names):
            for c, sB in enumerate(student_names):
                if sA == sB:
                    continue
                cell = ws.cell(row=r + row_off, column=c + col_off)
                cell.alignment = Alignment(horizontal='center')
                value, comment = scorer(sA, sB)
                cell.value = value
                if comment:
                    cell.comment = comment
        self._format_asymmetric_sheet(ws, student_names, has_ids)

    def generate_matrix_report(self, output_file: str) -> None:
        if not self.student_data:
            self.load_student_data()

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        def _sort_key(n):
            sid = self.id_map.get(n, '')
            try:
                return (0, int(sid), n)
            except (TypeError, ValueError):
                return (1, sid, n)

        student_names = sorted(self.student_data.keys(), key=_sort_key)
        active_exts = [ext for ext in self.extensions
                       if any(self.student_data[n].get(ext) is not None
                              for n in self.student_data)] or self.extensions

        print("Generating similarity matrices...")

        def _avg_score(sA, sB, fn):
            scores = [fn(self.student_data[sA][ext], self.student_data[sB][ext])
                      for ext in active_exts
                      if self.student_data[sA].get(ext) and self.student_data[sB].get(ext)]
            return (round(sum(scores) / len(scores), 2) or None) if scores else None

        def _score_diff(sA, sB):
            return _avg_score(sA, sB, lambda a, b: calculate_ide_diff_sim(a, b) * 100), None

        def _score_char(sA, sB):
            return _avg_score(sA, sB, lambda a, b: calculate_char_histogram_similarity(a, b) * 100), None

        def _score_inc(sA, sB):
            fA = self.student_outside_full.get(sA, Counter())
            fB = self.student_outside_full.get(sB, Counter())
            if fA and fB:
                return round(calculate_containment(fA, fB), 2) or None, None
            return None, None

        def _score_extra(sA, sB):
            eA = sum(self.student_extra_outside[sA].values(), Counter())
            eB = sum(self.student_extra_outside[sB].values(), Counter())
            overlap = sum((eA & eB).values())
            return overlap or None, _extra_comment(eA, eB, sA)

        def _score_extra_c(sA, sB):
            iA = sum(self.student_extra_inside[sA].values(), Counter())
            iB = sum(self.student_extra_inside[sB].values(), Counter())
            overlap = sum((iA & iB).values())
            return overlap or None, _extra_comment(iA, iB, sA)

        for title, scorer in [
            ('Diff',       _score_diff),
            ('Char',       _score_char),
            ('Inc',        _score_inc),
            ('Extra',      _score_extra),
            ('Extra (C)',  _score_extra_c),
        ]:
            self._fill_matrix_sheet(wb, title, student_names, scorer)

        save_xlsx(wb, output_file)
        print(f"Report saved to {output_file}")


    @staticmethod
    def _init_matrix_header(ws, student_names, student_ids=None):
        has_ids = student_ids is not None
        name_row = 2 if has_ids else 1
        name_col = 2 if has_ids else 1
        first_data_col = name_col + 1
        ws.row_dimensions[name_row].height = _HEADER_ROW_HEIGHT
        if has_ids:
            ws.row_dimensions[1].height = _HEADER_ROW_HEIGHT
        for idx, name in enumerate(student_names):
            data_col = idx + first_data_col
            c = ws.cell(row=name_row, column=data_col, value=name)
            c.font = Font(bold=True)
            c.alignment = Alignment(text_rotation=90, vertical='bottom',
                                    horizontal='center')
            ws.cell(row=idx + name_row + 1, column=name_col,
                    value=name).font = Font(bold=True)
            if has_ids:
                sid = student_ids[idx]
                ic = ws.cell(row=1, column=data_col, value=sid)
                ic.font = Font(bold=True)
                ic.alignment = Alignment(text_rotation=90, vertical='bottom',
                                         horizontal='center')
                ws.cell(row=idx + name_row + 1, column=1,
                        value=sid).font = Font(bold=True)

    @staticmethod
    def _format_asymmetric_sheet(ws, student_names, has_ids=False):
        n = len(student_names)
        name_row = 2 if has_ids else 1
        name_col = 2 if has_ids else 1
        first_data_col = name_col + 1
        first_data_row = name_row + 1
        data_col_ltr = ws.cell(row=name_row, column=first_data_col).column_letter
        max_col_idx  = first_data_col + n - 1
        max_col_ltr  = ws.cell(row=name_row, column=max_col_idx).column_letter
        max_row      = first_data_row + n - 1
        ws.freeze_panes = f'{data_col_ltr}{first_data_row}'
        ws.conditional_formatting.add(
            f'{data_col_ltr}{first_data_row}:{max_col_ltr}{max_row}',
            ColorScaleRule(start_type='num', start_value=0, start_color='FFFFFF',
                           end_type='max', end_color='F8696B'),
        )
        ws.column_dimensions['A'].width = 6 if has_ids else 20
        if has_ids:
            ws.column_dimensions[ws.cell(row=1, column=name_col).column_letter].width = 20
        for col_idx in range(first_data_col, max_col_idx + 1):
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 6
        for row_cells in ws.iter_rows(min_row=first_data_row, max_row=max_row,
                                      min_col=first_data_col, max_col=max_col_idx):
            for cell in row_cells:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0'
                    if cell.value >= 100:
                        cell.border = _BLACK_BORDER
                        cell.font   = Font(bold=True, color='FFFF00')


def main():
    if len(sys.argv) < 2:
        print('Usage: peer_sim_check.py <project_dir>'); sys.exit(1)
    current_dir  = Path(sys.argv[1]).resolve()
    anon_ids_dir = current_dir / 'anon_ids'
    correct_dir  = current_dir / 'correct'
    start_dir    = current_dir / 'start'

    if not correct_dir.exists():
        print(f"Missing: {correct_dir}"); sys.exit(1)

    if anon_ids_dir.exists():
        id_map = {d.name: d.name for d in anon_ids_dir.iterdir() if d.is_dir()}
        if id_map:
            print(f"Found {len(id_map)} student folder(s).")
        log_data, log_msg = load_lesson_log(current_dir)
        if log_msg:
            print(log_msg)
        events = log_data.all_events if log_data else None
        lesson_file = log_data.lesson_file if log_data else None
        checker = PeerSimilarityChecker(
            str(anon_ids_dir), str(correct_dir),
            start_dir=str(start_dir) if start_dir.exists() else None,
            id_map=id_map,
            events=events,
            lesson_file=lesson_file,
        )
        excels_dir = current_dir / 'excels'
        excels_dir.mkdir(exist_ok=True)
        checker.generate_matrix_report(
            str(excels_dir / 'student_similarity.xlsx')
        )
    else:
        print(f"Missing: {anon_ids_dir}")


if __name__ == "__main__":
    main()