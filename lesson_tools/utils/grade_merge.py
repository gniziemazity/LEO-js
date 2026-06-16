from pathlib import Path

from openpyxl import load_workbook

from .similarity_measures import save_xlsx


def merge_manual_columns(src_path: Path, dst_path: Path) -> None:
    try:
        src_wb = load_workbook(src_path)
        dst_wb = load_workbook(dst_path)
        sheet = 'Remarks'
        if sheet not in src_wb.sheetnames or sheet not in dst_wb.sheetnames:
            return
        src_ws, dst_ws = src_wb[sheet], dst_wb[sheet]

        def _hdr(ws, *names):
            return {c.value: c.column for c in ws[1] if c.value in names}

        src_c = _hdr(src_ws, 'Obs', 'Grade', 'Status', 'Comments')
        dst_c = _hdr(dst_ws, 'Obs', 'Grade', 'Status', 'Comments')
        if not (src_c and dst_c):
            return

        src_map = {
            str(row[0].value): {n: row[col - 1] for n, col in src_c.items()}
            for row in src_ws.iter_rows(min_row=2)
            if row[0].value is not None
        }
        for row in dst_ws.iter_rows(min_row=2):
            sid = str(row[0].value) if row[0].value is not None else None
            if sid not in src_map:
                continue
            for name, dst_col in dst_c.items():
                sc = src_map[sid].get(name)
                if sc is None:
                    continue
                dc = row[dst_col - 1]
                dc.value = sc.value
                if sc.has_style:
                    import copy

                    dc.font = copy.copy(sc.font)
                    dc.fill = copy.copy(sc.fill)
                    dc.border = copy.copy(sc.border)
                    dc.alignment = copy.copy(sc.alignment)
                    dc.number_format = sc.number_format

        save_xlsx(dst_wb, str(dst_path), vml_source=str(dst_path))
        print(f'  Merged Obs/Grade/Status/Comments from: {src_path.name}')
    except PermissionError:
        raise
    except Exception as e:
        print(f'  Warning: could not merge manual columns: {e}')
