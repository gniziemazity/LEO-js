import os
from collections import Counter
from pathlib import Path
from typing import Dict, List, NamedTuple, Tuple

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .folder_utils import LANG_EXTS
from .similarity_measures import (
    calculate_containment,
    iter_code_tokens,
    save_xlsx,
    token_edit_similarity,
)
from .token_log import _split_tokens_by_comment
from .token_log_mixin import (
    _LANG_EXT_LABEL,
    _effective_ext_at,
    _embedded_lang_ranges_for,
    _ext_of,
)


def _fmt_diff(miss_ctr: Counter, extra_ctr: Counter) -> Tuple[str, List[str]]:
    parts: List[str] = []
    items: List[str] = []
    for kw, n in sorted(miss_ctr.items()):
        suf = f' (x{n})' if n > 1 else ''
        parts.append(f'-{kw}{suf}')
        items.append(f'Missing: {kw}{suf}')
    for kw, n in sorted(extra_ctr.items()):
        suf = f' (x{n})' if n > 1 else ''
        parts.append(f'+{kw}{suf}')
        items.append(f'Extra: {kw}{suf}')
    return (', '.join(parts), items)


def _ts_str(seconds: int) -> str:
    h, rem = divmod(seconds, 3600)
    m, s = divmod(rem, 60)
    return f'{h:02d}:{m:02d}:{s:02d}'


def _build_synth_ts(files: Dict[str, Path]) -> Dict[Tuple[str, int], str]:
    out: Dict[Tuple[str, int], str] = {}
    counter = 0
    for _ext, fpath in sorted((files or {}).items(), key=lambda kv: kv[1].name):
        try:
            text = fpath.read_text(encoding='utf-8', errors='ignore')
        except Exception:
            continue
        file_ext = _ext_of(fpath.name)
        for pos, _tok, _is_comment in iter_code_tokens(text, file_ext):
            out[(fpath.name, pos)] = _ts_str(counter)
            counter += 1
    return out


class _Col(NamedTuple):
    key: str
    header: str
    hidden: bool = False
    cf: str = ''


_CF_RULES = {
    'extra': lambda: ColorScaleRule(
        start_type='num', start_value=0, start_color='FFFFFF',
        end_type='max', end_color='F8696B'),
    'inc': lambda: ColorScaleRule(
        start_type='num', start_value=0, start_color='F8696B',
        end_type='num', end_value=100, end_color='FFFFFF'),
    'redlow': lambda: ColorScaleRule(
        start_type='min', start_color='F8696B',
        end_type='max', end_color='FFFFFF'),
}

_OPTIONAL_TRAIL = ('status', 'comments', 'category')


class ExcelReportMixin:
    def generate_remarks_report(
        self,
        output_file: str,
        anonymize: bool = False,
        token_stats: 'Dict[str, dict] | None' = None,
        basis_marks_by_sid: 'Dict[str, dict] | None' = None,
    ) -> None:
        wb = Workbook()
        wb.remove(wb.active)
        prev_stats = None
        prev_basis = getattr(self, '_basis_marks_by_sid', None)
        if token_stats is not None:
            prev_stats = self._student_token_stats
            self._student_token_stats = token_stats
        if basis_marks_by_sid is not None:
            self._basis_marks_by_sid = basis_marks_by_sid
        try:
            self._add_remarks_sheet(wb, anonymize=anonymize)
            save_xlsx(wb, output_file)
        finally:
            if prev_stats is not None:
                self._student_token_stats = prev_stats
            self._basis_marks_by_sid = prev_basis

    def _add_remarks_sheet(self, wb: Workbook, anonymize: bool = False) -> None:
        sheet = wb.create_sheet(title='Remarks')
        is_assignment = not bool(self._lesson_keypresses)

        present_lang_exts = self._present_lang_exts(is_assignment)
        cols = self._remarks_columns(is_assignment, present_lang_exts)
        header = [c.header for c in cols]
        idx = {c.key: i + 1 for i, c in enumerate(cols)}

        sheet.append(header)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        excluded = getattr(self, 'excluded_ids', set()) or set()
        ai_ids = getattr(self, 'ai_ids', set()) or set()
        anon_mode = (
            os.environ.get('STUDENT_ANALYTICS_USE_ALTER_EGO') == '1'
            or anonymize
        )
        sids = sorted(self.student_info.keys(), key=int)
        active_sids = [s for s in sids if s not in excluded]
        all_extra_counters = {
            sid: sum(self.student_simple_extra_by_ext.get(sid, {}).values(), Counter())
            for sid in active_sids
        }
        _extra_denom = max(
            1, sum(sum(c.values()) for c in self.teacher_outside_by_ext.values())
        )
        student_interactions = self._extract_interactions() if not is_assignment else {}

        for sid in sids:
            info = self.student_info[sid]
            display_name = sid if anonymize else info['name']
            display_number = (
                '123456'
                if anon_mode and sid not in ai_ids
                else info['number']
            )
            if sid in excluded:
                vals = {'id': int(sid), 'student': display_name,
                        'number': display_number, 'category': 'EXCLUDED'}
                sheet.append(self._row_cells(cols, vals, full=True))
                continue

            extra_ctr = all_extra_counters.get(sid, Counter())
            extra_all = [f'{kw} (x{n})' if n > 1 else kw
                         for kw, n in sorted(extra_ctr.items())]
            has_submission = sid in self.results
            code_not_found = has_submission and not self.student_raw_texts.get(sid, '').strip()

            remarks_emoji, details = self._remarks_emoji(
                info['number'], has_submission, code_not_found
            )
            inc_pct = self._avg_inc(sid) if has_submission and not code_not_found else ''
            _ts = self._student_token_stats.get(sid)
            if _ts and has_submission and not code_not_found:
                follow_e_pct = _ts['follow_e']
                comment_pct = _ts['follow_c']
                ts_denom_r = _ts['teacher_total_e'] or 1
                extra_pct = (round(min(_ts['extra'] / ts_denom_r * 100, 100.0), 1)
                             if _ts['extra'] else 0.0)
                extra_e_text = _ts['extra_e_text']
                comment_text = _ts['comment_text']
                extra_all = _ts['extra_all']
            else:
                follow_e_pct = comment_pct = ''
                extra_e_text = comment_text = ''
                extra_pct = (round(min(sum(extra_ctr.values()) / _extra_denom * 100, 100.0), 1)
                             if has_submission and not code_not_found else '')

            req_count, req_details, req_missing, req_fill = self._check_required(
                sid, has_submission, code_not_found
            )
            sim_by_lang: Dict[str, Dict] = {}

            vals: Dict[str, object] = {
                'id': int(sid), 'student': display_name, 'number': display_number,
            }
            if code_not_found:
                vals['remarks'] = '⛔'
            else:
                vals['remarks'] = remarks_emoji
                vals['extra'] = extra_pct
                vals['extra_t'] = ', '.join(extra_all) if extra_all else ''
                vals['inc'] = inc_pct
                if not is_assignment:
                    vals['follow_c'] = comment_pct
                    vals['follow_c_t'] = comment_text
                    vals['follow_e'] = follow_e_pct
                    vals['follow_e_t'] = extra_e_text
                    vals['interact'] = student_interactions.get(sid, '')
                    lang_scores = (_ts or {}).get('follow_e_by_lang') or {}
                    for ext in present_lang_exts:
                        v = lang_scores.get(ext)
                        if v is not None:
                            vals[f'lang:{ext}'] = v.get('score', '')
                            vals[f'lang_t:{ext}'] = v.get('text', '')
                else:
                    basis_marks = (getattr(self, '_basis_marks_by_sid', None) or {}).get(sid)
                    pb_info = (
                        self._per_basis_sim_info(sid, basis_marks)
                        if basis_marks else None
                    )
                    if pb_info is not None:
                        sim_pct, sim_desc, _sim_items = pb_info
                    else:
                        sim_pct, sim_desc, _sim_items = self._similarity_info(
                            sid, has_submission, code_not_found)
                    vals['sim'] = sim_pct
                    vals['sim_t'] = sim_desc
                    sim_by_lang = (
                        self._similarity_info_by_lang_from_marks(
                            sid, has_submission, code_not_found, basis_marks)
                        if basis_marks
                        else self._similarity_info_by_lang(
                            sid, has_submission, code_not_found)
                    )
                    for ext in present_lang_exts:
                        v = sim_by_lang.get(ext)
                        if v is not None:
                            vals[f'lang:{ext}'] = v['score']
                            vals[f'lang_t:{ext}'] = v['desc']
                if self.required_items or self.not_expected_items:
                    vals['expected'] = req_count
                    vals['expected_t'] = req_details
                vals['obs'] = '_' if has_submission else ''
            if sid in ai_ids:
                vals['category'] = 'LLM'

            sheet.append(self._row_cells(cols, vals, full=code_not_found))
            cur = sheet.max_row

            if req_fill and 'expected' in idx:
                sheet.cell(row=cur, column=idx['expected']).fill = req_fill

            rem_col = idx['remarks']
            if code_not_found:
                c = Comment('No code files found in submission', 'sim_check')
                c.width = 400; c.height = 80
                sheet.cell(row=cur, column=rem_col).comment = c
                sheet.cell(row=cur, column=rem_col).font = Font(name='Segoe UI Emoji')
            elif details:
                items = [r for r in details.split('; ') if r]
                c = Comment(', '.join(items), 'sim_check')
                c.width = 500; c.height = min(100 + 30 * len(items), 1200)
                sheet.cell(row=cur, column=rem_col).comment = c
            else:
                sheet.cell(row=cur, column=rem_col).font = Font(name='Segoe UI Emoji')

            if extra_all and not code_not_found:
                c = Comment(', '.join(extra_all), 'sim_check')
                c.width = 400
                c.height = min(200 + 80 * len(extra_all), 6000)
                sheet.cell(row=cur, column=idx['extra']).comment = c

            if not is_assignment and comment_text:
                _c_items = (_ts['comment_items'] if _ts and _ts.get('comment_items')
                            else comment_text.split(', '))
                c = Comment(', '.join(_c_items), 'sim_check')
                c.width = 400
                c.height = min(200 + 80 * len(_c_items), 6000)
                sheet.cell(row=cur, column=idx['follow_c']).comment = c

            if not is_assignment and extra_e_text:
                _e_items = (_ts['extra_e_items'] if _ts and _ts.get('extra_e_items')
                            else extra_e_text.split(', '))
                c = Comment(', '.join(_e_items), 'sim_check')
                c.width = 400
                c.height = min(200 + 80 * len(_e_items), 6000)
                sheet.cell(row=cur, column=idx['follow_e']).comment = c

            if not is_assignment and not code_not_found and _ts:
                _lang_scores = _ts.get('follow_e_by_lang') or {}
                for ext in present_lang_exts:
                    lang_v = _lang_scores.get(ext)
                    if not lang_v:
                        continue
                    text = lang_v.get('text') or ''
                    if not text:
                        continue
                    items = lang_v.get('items') or [text]
                    c = Comment(text, 'sim_check')
                    c.width = 400
                    c.height = min(200 + 80 * len(items), 6000)
                    sheet.cell(row=cur, column=idx[f'lang:{ext}']).comment = c

            if is_assignment and not code_not_found and sim_by_lang:
                for ext in present_lang_exts:
                    v = sim_by_lang.get(ext)
                    if not v:
                        continue
                    desc = v.get('desc') or ''
                    if not desc:
                        continue
                    items = v.get('items') or [desc]
                    c = Comment(desc, 'sim_check')
                    c.width = 400
                    c.height = min(200 + 80 * len(items), 6000)
                    sheet.cell(row=cur, column=idx[f'lang:{ext}']).comment = c

            if is_assignment and not code_not_found and sim_desc and 'sim' in idx:
                c = Comment(sim_desc, 'sim_check')
                c.width = 400
                c.height = 200
                sheet.cell(row=cur, column=idx['sim']).comment = c

            if (self.required_items or self.not_expected_items) and req_missing and 'expected' in idx:
                c = Comment(', '.join(req_missing), 'sim_check')
                c.width = 500; c.height = min(100 + 30 * len(req_missing), 1200)
                sheet.cell(row=cur, column=idx['expected']).comment = c

        self._apply_remarks_formatting(sheet, cols, idx)

    @staticmethod
    def _row_cells(cols: 'List[_Col]', vals: Dict[str, object],
                   full: bool) -> List[object]:
        out: List[object] = []
        for c in cols:
            if c.key in vals:
                out.append(vals[c.key])
            elif full or c.key not in _OPTIONAL_TRAIL:
                out.append('')
            else:
                out.append(None)
        return out

    def _present_lang_exts(self, is_assignment: bool) -> List[str]:
        lang_exts = [ext for ext, _label in _LANG_EXT_LABEL]
        if not is_assignment:
            return [
                ext for ext in lang_exts
                if any(
                    (v.get('follow_e_by_lang') or {}).get(ext) is not None
                    for v in self._student_token_stats.values()
                )
            ]
        teacher_by_lang = self._teacher_tokens_by_lang()
        return [ext for ext in lang_exts if teacher_by_lang.get(ext)]

    def _remarks_columns(self, is_assignment: bool,
                         present_lang_exts: List[str]) -> 'List[_Col]':
        lang_label = {ext: f'{label} (E)' for ext, label in _LANG_EXT_LABEL}
        cols = [
            _Col('id', 'ID'),
            _Col('student', 'Student'),
            _Col('number', 'Number'),
            _Col('remarks', 'Remarks'),
            _Col('extra', 'Extra', cf='extra'),
            _Col('extra_t', 'Extra Desc', hidden=True),
            _Col('inc', 'Inc', cf='inc'),
        ]
        if not is_assignment:
            cols += [
                _Col('follow_c', 'Follow (C)', cf='redlow'),
                _Col('follow_c_t', 'Follow (C) Desc', hidden=True),
                _Col('follow_e', 'Follow (E)', cf='redlow'),
                _Col('follow_e_t', 'Follow (E) Desc', hidden=True),
                _Col('interact', 'Interactions'),
            ]
        else:
            cols += [
                _Col('sim', 'Similarity', cf='redlow'),
                _Col('sim_t', 'Similarity Desc', hidden=True, cf='redlow'),
            ]
        for ext in present_lang_exts:
            cols.append(_Col(f'lang:{ext}', lang_label[ext], cf='redlow'))
            cols.append(_Col(f'lang_t:{ext}', f'{lang_label[ext]} Desc', hidden=True))
        if self.required_items or self.not_expected_items:
            cols += [
                _Col('expected', 'Expected'),
                _Col('expected_t', 'Expected Desc', hidden=bool(self.required_items)),
            ]
        cols += [_Col('obs', 'Obs'), _Col('grade', 'Grade')]
        if is_assignment:
            cols.append(_Col('status', 'Status'))
        cols += [_Col('comments', 'Comments'), _Col('category', 'Category')]
        return cols

    def _apply_remarks_formatting(self, sheet, cols: 'List[_Col]',
                                  idx: Dict[str, int]) -> None:
        for c in cols:
            if c.hidden:
                sheet.column_dimensions[get_column_letter(idx[c.key])].hidden = True

        max_row = sheet.max_row
        GREEN_FILL = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        RED_FILL = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        rem_col = idx['remarks']
        for r in range(2, max_row + 1):
            val = sheet.cell(row=r, column=rem_col).value
            if val == '✅':
                sheet.cell(row=r, column=rem_col).fill = GREEN_FILL
            elif val == '⛔':
                sheet.cell(row=r, column=rem_col).fill = RED_FILL

        if max_row > 1:
            for c in cols:
                if not c.cf:
                    continue
                ltr = get_column_letter(idx[c.key])
                sheet.conditional_formatting.add(
                    f'{ltr}2:{ltr}{max_row}', _CF_RULES[c.cf]())

        self._auto_column_widths(sheet)
        sheet.column_dimensions['B'].width = 18

    def _similarity_info(self, sid: str, has_submission: bool, code_not_found: bool):
        if not has_submission or code_not_found:
            return ('', '', [])
        data = self.results.get(sid, {})
        if not data or not data.get('files_compared'):
            return ('', '', [])
        inc_vals    = []
        teacher_agg: Counter = Counter()
        for ext in LANG_EXTS:
            fd = data['files_compared'].get(ext)
            if not fd or fd.get('status') != 'success':
                continue
            inc_vals.append(fd['inc_sim'])
            if ext == '.html':
                teacher_ext = self.teacher_outside_by_ext.get(ext, Counter())
            elif ext == '.css':
                teacher_ext = self.teacher_tokens_by_ext.get(ext, Counter())
            else:
                teacher_ext = self.teacher_tokens_by_ext.get(ext, Counter())
            teacher_agg += teacher_ext
        student_agg: Counter = getattr(self, '_student_all_outside', {}).get(sid, Counter())

        sim_desc, sim_items = _fmt_diff(teacher_agg - student_agg, student_agg - teacher_agg)
        sim_pct = round(sum(inc_vals) / len(inc_vals), 1) if inc_vals else ''
        return (sim_pct, sim_desc, sim_items)

    def _build_synth_teacher_timestamps(self) -> None:
        self._synth_ts_teacher = _build_synth_ts(
            self.get_code_files(self._effective_reference_dir())
        )

    def _per_basis_sim_info(self, sid: str, basis_marks: dict):
        if not basis_marks:
            return None
        miss_marks: List[Tuple[str, int, str]] = []
        extra_marks: List[Tuple[str, int, str]] = []
        n_extra_unpaired = 0
        n_ghost_extra = 0
        for fname, marks in (basis_marks.get('teacher_files') or {}).items():
            for m in marks or []:
                if m.get('label') == 'missing':
                    miss_marks.append((fname, m.get('start', 0), m.get('token', '')))
        for fname, marks in (basis_marks.get('student_files') or {}).items():
            for m in marks or []:
                lbl = m.get('label')
                if lbl not in ('extra', 'ghost_extra'):
                    continue
                extra_marks.append((fname, m.get('start', 0), m.get('token', '')))
                if lbl == 'ghost_extra':
                    n_ghost_extra += 1
                elif not m.get('paired_with'):
                    n_extra_unpaired += 1
        teacher_by_lang = self._teacher_tokens_by_lang()
        teacher_total = sum(sum(ctr.values()) for ctr in teacher_by_lang.values())
        if not teacher_total:
            return None
        deduction = len(miss_marks) + n_extra_unpaired + n_ghost_extra
        score = round(max(0.0, (teacher_total - deduction) / teacher_total * 100), 1)
        ts_teacher = getattr(self, '_synth_ts_teacher', None) or {}
        miss_sorted = sorted(
            ((ts_teacher.get((fn, s), '99:99:99'), tok)
             for fn, s, tok in miss_marks),
        )
        extras_sorted = sorted(extra_marks, key=lambda t: (t[0], t[1]))
        parts: List[str] = []
        items: List[str] = []
        for ts, tok in miss_sorted:
            ts_s = '' if ts == '99:99:99' else f' ({ts})'
            parts.append(f'-{tok}{ts_s}')
            items.append(f'Missing: {tok}{ts_s}')
        for _fname, _pos, tok in extras_sorted:
            parts.append(f'+{tok} (00:00:00)')
            items.append(f'Extra: {tok}')
        desc = ', '.join(parts)
        return (score, desc, items)

    def _similarity_info_by_lang_from_marks(
        self, sid: str, has_submission: bool, code_not_found: bool, basis_marks: dict,
    ) -> Dict[str, Dict]:
        if not has_submission or code_not_found or not basis_marks:
            return {}
        teacher_by_lang = self._teacher_tokens_by_lang()
        teacher_files = self.get_code_files(self._effective_reference_dir())
        teacher_texts: Dict[str, str] = {}
        teacher_ranges: Dict[str, Dict[str, List[Tuple[int, int]]]] = {}
        for ext, fpath in (teacher_files or {}).items():
            try:
                text = fpath.read_text(encoding='utf-8', errors='ignore')
            except Exception:
                continue
            teacher_texts[fpath.name] = text
            teacher_ranges[fpath.name] = _embedded_lang_ranges_for(text, ext)
        student_dir = getattr(self, 'student_dir_by_sid', {}).get(sid)
        student_files = self.get_code_files(student_dir) if student_dir else {}
        student_ranges: Dict[str, Dict[str, List[Tuple[int, int]]]] = {}
        for ext, fpath in (student_files or {}).items():
            try:
                text = fpath.read_text(encoding='utf-8', errors='ignore')
            except Exception:
                continue
            student_ranges[fpath.name] = _embedded_lang_ranges_for(text, ext)


        miss_by_lang: Dict[str, List[Tuple[str, int, str]]] = {}
        extra_by_lang: Dict[str, List[Tuple[str, int, str]]] = {}
        n_extra_unpaired_by_lang: Dict[str, int] = {}
        n_ghost_extra_by_lang: Dict[str, int] = {}
        for fname, marks in (basis_marks.get('teacher_files') or {}).items():
            file_ext = _ext_of(fname)
            if not file_ext:
                continue
            ranges = teacher_ranges.get(fname, {})
            for m in marks or []:
                if m.get('label') != 'missing':
                    continue
                pos = m.get('start', 0)
                eff = _effective_ext_at(pos, file_ext, ranges) if ranges else file_ext
                pw = m.get('paired_with')
                repl = pw.get('token', '') if pw else None
                miss_by_lang.setdefault(eff, []).append((fname, pos, m.get('token', ''), repl))
        for fname, marks in (basis_marks.get('student_files') or {}).items():
            file_ext = _ext_of(fname)
            if not file_ext:
                continue
            ranges = student_ranges.get(fname, {})
            for m in marks or []:
                lbl = m.get('label')
                if lbl not in ('extra', 'ghost_extra'):
                    continue
                pos = m.get('start', 0)
                eff = _effective_ext_at(pos, file_ext, ranges) if ranges else file_ext
                extra_by_lang.setdefault(eff, []).append((fname, pos, m.get('token', '')))
                if lbl == 'ghost_extra':
                    n_ghost_extra_by_lang[eff] = n_ghost_extra_by_lang.get(eff, 0) + 1
                elif not m.get('paired_with'):
                    n_extra_unpaired_by_lang[eff] = n_extra_unpaired_by_lang.get(eff, 0) + 1

        ts_teacher = getattr(self, '_synth_ts_teacher', None) or {}

        out: Dict[str, Dict] = {}
        for ext in LANG_EXTS:
            teacher_ext_total = sum(teacher_by_lang.get(ext, Counter()).values())
            miss = miss_by_lang.get(ext, [])
            extra = extra_by_lang.get(ext, [])
            if not teacher_ext_total and not miss and not extra:
                continue
            n_miss = len(miss)
            n_extra_unpaired = n_extra_unpaired_by_lang.get(ext, 0)
            n_ghost_extra = n_ghost_extra_by_lang.get(ext, 0)
            deduction = n_miss + n_extra_unpaired + n_ghost_extra
            score = (
                round(max(0.0, (teacher_ext_total - deduction) / teacher_ext_total * 100), 1)
                if teacher_ext_total else 0.0
            )
            miss_sorted = sorted(
                ((ts_teacher.get((fn, s), '99:99:99'), tok, repl)
                 for fn, s, tok, repl in miss),
                key=lambda x: (x[0], x[1]),
            )
            extras_sorted = sorted(extra, key=lambda t: (t[0], t[1]))
            parts: List[str] = []
            items: List[str] = []
            for ts, tok, repl in miss_sorted:
                ts_s = '' if ts == '99:99:99' else f' ({ts})'
                sim_s = f' ~{token_edit_similarity(tok, repl):.2f}' if repl is not None else ''
                parts.append(f'-{tok}{ts_s}{sim_s}')
                items.append(f'Missing: {tok}{ts_s}')
            for _fname, _pos, tok in extras_sorted:
                parts.append(f'+{tok} (00:00:00)')
                items.append(f'Extra: {tok}')
            desc = ', '.join(parts)
            out[ext] = {
                'score': score,
                'desc':  desc,
                'items': items,
            }
        return out

    def _tokens_by_effective_lang(self, files: Dict[str, Path]) -> Dict[str, Counter]:
        by_lang: Dict[str, Counter] = {}
        for ext, file_path in (files or {}).items():
            try:
                text = file_path.read_text(encoding='utf-8', errors='ignore')
            except Exception:
                continue
            ranges = _embedded_lang_ranges_for(text, ext)
            nc, _cm = _split_tokens_by_comment(text, ext)
            for pos, tok in nc:
                eff = _effective_ext_at(pos, ext, ranges) if ranges else ext
                by_lang.setdefault(eff, Counter())[tok] += 1
        return by_lang

    def _teacher_tokens_by_lang(self) -> Dict[str, Counter]:
        cache = getattr(self, '_teacher_by_lang_cache', None)
        if cache is None:
            cache = self._tokens_by_effective_lang(
                self.get_code_files(self._effective_reference_dir())
            )
            self._teacher_by_lang_cache = cache
        return cache

    def _student_tokens_by_lang(self, sid: str) -> Dict[str, Counter]:
        student_dir = getattr(self, 'student_dir_by_sid', {}).get(sid)
        if student_dir is None:
            return {}
        return self._tokens_by_effective_lang(self.get_code_files(student_dir))

    def _similarity_info_by_lang(self, sid: str, has_submission: bool, code_not_found: bool) -> Dict[str, Dict]:
        if not has_submission or code_not_found:
            return {}
        teacher_by_lang = self._teacher_tokens_by_lang()
        student_by_lang = self._student_tokens_by_lang(sid)
        out: Dict[str, Dict] = {}
        for ext in LANG_EXTS:
            teacher_ext = teacher_by_lang.get(ext, Counter())
            student_ext = student_by_lang.get(ext, Counter())
            if not teacher_ext and not student_ext:
                continue
            score = calculate_containment(teacher_ext, student_ext)
            desc, items = _fmt_diff(teacher_ext - student_ext, student_ext - teacher_ext)
            out[ext] = {
                'score': score,
                'desc':  desc,
                'items': items,
            }
        return out

    def _avg_inc(self, sid: str) -> str:
        inc_vals = [fd['inc_sim']
                    for fd in self.results[sid]['files_compared'].values()
                    if fd.get('status') == 'success']
        return round(sum(inc_vals) / len(inc_vals), 1) if inc_vals else ''

    def _remarks_emoji(self, student_number: str, has_submission: bool, code_not_found: bool):
        if not has_submission:
            return '', ''
        if code_not_found:
            return '⛔', ''
        raw = self.remarks_data.get(student_number, '')
        if not raw or raw == 'OK':
            return '✅', ''
        items     = [r for r in raw.split('; ') if r]
        emoji_map = []
        for item in items:
            il = item.lower()
            if 'student number not found' in il:
                emoji_map.append('🪪')
            elif 'redacted' in il:
                emoji_map.append('◼️')
            elif 'another student number' in il or 'unknown digit sequence' in il:
                emoji_map.append('🤥')
            else:
                emoji_map.append('⚠️')
        seen = set()
        deduped = []
        for e in emoji_map:
            if e not in seen:
                seen.add(e); deduped.append(e)
        return ''.join(deduped), raw

    def _extract_interactions(self) -> Dict[str, str]:
        def _ids_from(raw) -> List[str]:
            if isinstance(raw, bool):
                return []
            if isinstance(raw, int):
                return [str(raw)]
            if isinstance(raw, list):
                return [str(x) for x in raw
                        if isinstance(x, int) and not isinstance(x, bool)]
            return []

        field_by_type = {
            'teacher-question': ('answered_by', 'A'),
            'student-question': ('asked_by', 'Q'),
            'providing-help':   ('student', 'H'),
        }
        iacts: List[Tuple[float, str, str]] = []
        for ev in self._lesson_interactions:
            spec = field_by_type.get(ev.get('interaction', ''))
            if spec is None:
                continue
            field_name, letter = spec
            ts_s = ev.get('timestamp', 0) / 1000
            for sid in _ids_from(ev.get(field_name)):
                iacts.append((ts_s, letter, sid))

        iacts.sort(key=lambda x: x[0])
        per_sid: Dict[str, List[str]] = {}
        for _, letter, sid in iacts:
            if sid in self.student_info:
                per_sid.setdefault(sid, []).append(letter)
        return {s: ', '.join(v) for s, v in per_sid.items()}

    def _check_required(self, sid: str, has_submission: bool, code_not_found: bool):
        if not (self.required_items or self.not_expected_items) \
                or not has_submission or code_not_found:
            return '', '', [], None

        raw_text       = self.student_raw_texts.get(sid, '').lower()
        raw_no_space   = raw_text.replace(' ', '')
        missing = [
            ' / '.join(alts)
            for alts in self.required_items
            if not any(
                a.lower() in raw_text or a.lower().replace(' ', '') in raw_no_space
                for a in alts
            )
        ]
        forbidden = [
            ' / '.join(alts)
            for alts in self.not_expected_items
            if any(
                a.lower() in raw_text or a.lower().replace(' ', '') in raw_no_space
                for a in alts
            )
        ]
        total      = len(self.required_items)
        found_cnt  = total - len(missing)
        count_parts = []
        if self.required_items:
            count_parts.append(f'{found_cnt}/{total}')
        if forbidden:
            count_parts.append(f'!{len(forbidden)}')
        req_count   = ', '.join(count_parts)
        req_missing = missing + [f'!{kw}' for kw in forbidden]
        req_details = ', '.join(req_missing)

        req_fill = None
        if total > 0 and found_cnt < total:
            intensity = 1.0 - found_cnt / total
            gb = int(255 - intensity * 148)
            req_fill = PatternFill(start_color=f'FF{gb:02X}{gb:02X}',
                                   end_color=f'FF{gb:02X}{gb:02X}', fill_type='solid')
        elif forbidden:
            req_fill = PatternFill(start_color='FFDD88', end_color='FFDD88', fill_type='solid')
        return req_count, req_details, req_missing, req_fill

    @staticmethod
    def _auto_column_widths(sheet) -> None:
        for column in sheet.columns:
            col_letter = column[0].column_letter
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            sheet.column_dimensions[col_letter].width = min(max_length + 2, 50)
