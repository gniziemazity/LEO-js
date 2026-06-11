import json
import re
from datetime import datetime, time

from lv_timeline_config import Config

_MIDNIGHT = time(0, 0, 0)


def normalize_names(value):
    if not value:
        return []
    if isinstance(value, list):
        return [str(s).strip() for s in value if str(s).strip()]
    return [p.strip() for p in str(value).split(',') if p.strip()]


def load_keypress_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    return data


def extract_interactions(key_presses):
    interactions = {
        'teacher-question': [],
        'student-question': [],
        'providing-help': []
    }
    for kp in key_presses:
        if 'interaction' in kp:
            itype = kp['interaction']
            if itype in interactions:
                entry = {
                    'timestamp': kp['timestamp'] / 1000,
                    'info': kp.get('info', ''),
                    'asked_by': normalize_names(kp.get('asked_by')),
                    'answered_by': normalize_names(kp.get('answered_by')),
                    'student': normalize_names(kp.get('student')),
                    'closed_at': kp.get('closed_at', None),
                }
                interactions[itype].append(entry)
    return interactions


def calculate_intervals(key_presses):
    intervals = []
    for i in range(1, len(key_presses)):
        intervals.append((key_presses[i]['timestamp'] - key_presses[i-1]['timestamp']) / 1000)
    return intervals


def analyze_typing_blocks(key_presses):
    if not key_presses:
        return [], [], [], []

    char_presses = [kp for kp in key_presses if 'char' in kp]
    if not char_presses:
        return [], [], [], []

    blocks = []
    current_block = [char_presses[0]]

    for i in range(1, len(char_presses)):
        time_gap = (char_presses[i]['timestamp'] - char_presses[i-1]['timestamp']) / 1000

        if time_gap < Config.BURST_GAP_THRESHOLD:
            current_block.append(char_presses[i])
        else:
            if len(current_block) >= Config.MIN_BURST_CHARS:
                blocks.append(current_block)
            current_block = [char_presses[i]]

    if len(current_block) >= Config.MIN_BURST_CHARS:
        blocks.append(current_block)

    timestamps = []
    rates = []
    texts = []
    widths = []

    _DELETE_CHARS_SET = {'\u21a2', '\u21a3', '\u26d4', '\u232b', '\u2326'}

    for block in blocks:
        start_time = block[0]['timestamp'] / 1000
        end_time = block[-1]['timestamp'] / 1000
        duration = end_time - start_time

        if duration == 0:
            duration = 1

        char_count = len(block)
        rate = (char_count / duration) * 60
        text = "".join([kp.get('char', '') for kp in block])

        center_time = (start_time + end_time) / 2
        width_days = duration / (24 * 3600)

        has_dev    = any(kp.get('_editor') == 'dev' for kp in block)
        has_remove = any(kp.get('char', '') in _DELETE_CHARS_SET for kp in block)

        timestamps.append(center_time)
        rates.append(rate)
        texts.append(text)
        widths.append(width_days)

    return timestamps, rates, texts, widths


def _parse_all_events_from_follow(follow_text, session_date):
    events = []
    for m in re.finditer(r'([^\(,]+?)\s*\((\d{2}:\d{2}:\d{2}(?:\.\d{1,3})?)\)', follow_text):
        label = m.group(1).strip()
        ts = m.group(2)
        fmt = '%H:%M:%S.%f' if '.' in ts else '%H:%M:%S'
        try:
            t = datetime.strptime(ts, fmt).time()
            events.append((label, datetime.combine(session_date, t)))
        except ValueError:
            pass
    return events


def _is_mistake_event(label):
    label = (label or '').strip()
    if label.startswith('-'):
        return True
    if label.startswith('+'):
        return label.rstrip().endswith('*')
    return False


def _positionable_event(label, dt):
    return dt is not None and dt.time() != _MIDNIGHT and _is_mistake_event(label)


def load_student_data_from_xlsx(remarks_path, session_start_ts, session_end_ts=None):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('openpyxl not available – cannot load student xlsx data.')
        return [], {}

    session_dt   = datetime.fromtimestamp(session_start_ts)
    session_date = session_dt.date()

    follow_data = {}
    id_to_name = {}
    try:
        wb_r = load_workbook(remarks_path, read_only=True, data_only=True)
        ws_r = wb_r['Remarks']
        rows_r = list(ws_r.iter_rows(values_only=True))
        wb_r.close()
        header_r      = list(rows_r[0]) if rows_r else []
        name_col_r    = None
        id_col_r      = None
        follow_num_col  = None
        follow_text_col = None
        for ci, h in enumerate(header_r):
            if h == 'Student' and name_col_r is None:
                name_col_r = ci
            elif h == 'ID' and id_col_r is None:
                id_col_r = ci
            elif h == 'Follow (E)' and follow_num_col is None:
                follow_num_col = ci
            elif h == 'Follow (E) Desc' and follow_text_col is None:
                follow_text_col = ci
        if name_col_r is None or follow_text_col is None:
            print('  Warning: could not locate Student / Follow (E) Desc columns in remarks.xlsx')
        else:
            for row in rows_r[1:]:
                name = str(row[name_col_r]).strip() if row[name_col_r] else ''
                if name and id_col_r is not None and row[id_col_r] is not None:
                    id_to_name[str(row[id_col_r]).strip()] = name
                ft   = row[follow_text_col]
                ft_str = str(ft).strip() if ft else ''
                fn   = row[follow_num_col] if follow_num_col is not None else None
                try:
                    fn_val = float(fn) if fn is not None and fn != '' else None
                except (TypeError, ValueError):
                    fn_val = None
                if name:
                    follow_data[name] = {
                        'events': (_parse_all_events_from_follow(ft_str, session_date) if ft_str else []),
                        'pct': fn_val,
                    }
    except Exception as e:
        print(f'  Error reading remarks.xlsx: {e}')

    students = []
    session_end_dt = datetime.fromtimestamp(session_end_ts) if session_end_ts else None
    for name in sorted(follow_data.keys()):
        fd = follow_data.get(name, {})
        follow_pct = fd.get('pct') if isinstance(fd, dict) else None
        if follow_pct is None:
            continue
        events = fd.get('events', []) if isinstance(fd, dict) else fd
        mistakes = [e for e in events if _positionable_event(e[0], e[1])]
        if mistakes:
            follow_dt = min(mistakes, key=lambda e: e[1])[1]
        elif session_end_dt is not None:
            follow_dt = session_end_dt
        else:
            continue
        students.append({
            'name': name,
            'follow_dt': follow_dt,
            'follow_events': events,
            'follow_pct': follow_pct,
        })
    return students, id_to_name
